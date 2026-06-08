# SPDX-License-Identifier: MIT
"""Deterministic builder for the COMPLEX synthetic XLSX fixture.

Produces ``tests/fixtures/complex/acme_complex.xlsx``: a 100% synthetic
(Acme-style, never proprietary) workbook that stresses the brand-xlsx engine
across as many Excel component types as openpyxl can author:

  * MULTIPLE sheets in a deliberate tab order (Cover, Inputs, Model, Summary,
    Data) - exercises multi-sheet structure + skeleton ordering.
  * NAMED regions of every geometry the extractor distinguishes:
      - a single-cell title anchor sitting under a MERGED header row,
      - a single-cell that sits in a FROZEN header band,
      - a multi-cell INPUTS block (sample-data candidate),
      - a multi-cell DATA region that is the BODY of a native table object,
      - a single named cell that is a cross-sheet formula target.
  * FORMULAS: in-sheet SUM/total rows, percent-of-total ratios, a cross-sheet
    reference (Summary pulls from Model and Inputs), and a SUBTOTAL - all
    authored verbatim so we can later assert they survive generation byte-exact.
  * NUMBER FORMATS: currency (accounting), percent, thousands, and ISO date.
  * A native TABLE object (``AcmeDataTbl``) with a banded table style.
  * CONDITIONAL FORMATTING: a color scale, a cell-is rule, and a formula rule.
  * FROZEN PANES on the data sheets.
  * A synthetic LOGO image placed in a worksheet header (drawing) - generated
    in-process (no external/proprietary asset on disk).
  * Named cell STYLES (``AcmeTitle``, ``AcmeHeader``, ``AcmeCurrency``,
    ``AcmePercent``, ``AcmeInput``) registered on the workbook and applied.
  * Demo / sample data rows the generator is expected to clear.

The output is content-reproducible within a fixed library set: a fixed timestamp
/ fixed image bytes / sorted defined names mean two rebuilds in the SAME
environment are identical. The committed binary is the source of truth; rebuilds
may differ byte-for-byte (even in size) across openpyxl versions (benign
serialization noise), so equality is asserted STRUCTURALLY, not by raw bytes
(see tests/test_fixture_determinism.py).

Run:
    PYTHONPATH=scripts .venv/bin/python tests/fixtures/builders/build_complex_xlsx.py
"""

from __future__ import annotations

import struct
import zlib
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(__file__).resolve().parents[1] / "complex" / "acme_complex.xlsx"

# Synthetic Acme brand palette (made-up; never proprietary).
ACME_NAVY = "FF1F3864"
ACME_TEAL = "FF2E8B8B"
ACME_AMBER = "FFE8A33D"
ACME_LIGHT = "FFEAF0F6"
WHITE = "FFFFFFFF"


# ---------------------------------------------------------------------------
# A tiny synthetic PNG logo generated in-process (no external asset).
# ---------------------------------------------------------------------------
def _synthetic_logo_png() -> bytes:
    """Return bytes of a deterministic 64x24 RGBA PNG (an 'Acme' navy block).

    Hand-built so the fixture carries a real <xdr:pic> drawing in a header
    without committing any external/proprietary image. Pixels are a navy field
    with an amber stripe - purely decorative brand mark.
    """
    w, h = 64, 24
    navy = (31, 56, 100, 255)
    amber = (232, 163, 61, 255)
    raw = bytearray()
    for y in range(h):
        raw.append(0)  # PNG filter type 0 (None) per scanline
        for x in range(w):
            px = amber if (8 <= y < 16 and 4 <= x < 60) else navy
            raw.extend(px)

    def _chunk(tag: bytes, data: bytes) -> bytes:
        return (
            struct.pack(">I", len(data))
            + tag
            + data
            + struct.pack(">I", zlib.crc32(tag + data) & 0xFFFFFFFF)
        )

    sig = b"\x89PNG\r\n\x1a\n"
    ihdr = struct.pack(">IIBBBBB", w, h, 8, 6, 0, 0, 0)  # 8-bit RGBA
    idat = zlib.compress(bytes(raw), 9)
    return sig + _chunk(b"IHDR", ihdr) + _chunk(b"IDAT", idat) + _chunk(b"IEND", b"")


# ---------------------------------------------------------------------------
# Named cell styles (registered once on the workbook).
# ---------------------------------------------------------------------------
def _register_named_styles(wb: Workbook) -> None:
    thin = Side(style="thin", color=ACME_NAVY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title = NamedStyle(name="AcmeTitle")
    title.font = Font(name="Arial", size=20, bold=True, color=ACME_NAVY)
    title.alignment = Alignment(horizontal="center", vertical="center")

    header = NamedStyle(name="AcmeHeader")
    header.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    header.fill = PatternFill("solid", fgColor=ACME_NAVY)
    header.alignment = Alignment(horizontal="center", vertical="center")
    header.border = border

    currency = NamedStyle(name="AcmeCurrency")
    currency.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    currency.font = Font(name="Calibri", size=11, color=ACME_NAVY)
    currency.border = border

    percent = NamedStyle(name="AcmePercent")
    percent.number_format = "0.0%"
    percent.font = Font(name="Calibri", size=11, color=ACME_TEAL)
    percent.border = border

    inp = NamedStyle(name="AcmeInput")
    inp.fill = PatternFill("solid", fgColor=ACME_LIGHT)
    inp.font = Font(name="Calibri", size=11, color=ACME_NAVY)
    inp.border = border

    for style in (title, header, currency, percent, inp):
        wb.add_named_style(style)


# ---------------------------------------------------------------------------
# Sheet builders.
# ---------------------------------------------------------------------------
def _build_cover(wb: Workbook) -> None:
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    # Merged title band A1:E1 with a single-cell named title anchor at A1.
    ws.merge_cells("A1:E1")
    ws["A1"] = "{{report_title}}"
    ws["A1"].style = "AcmeTitle"
    ws.merge_cells("A2:E2")
    ws["A2"] = "{{report_subtitle}}"
    ws["A2"].font = Font(name="Arial", size=12, italic=True, color=ACME_TEAL)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A4"] = "Prepared for"
    ws["B4"] = "{{client_name}}"
    ws["A5"] = "Reporting period"
    ws["B5"] = "{{period}}"
    # An ISO-date cell with a date number format.
    ws["A6"] = "Generated on"
    ws["B6"] = "2026-01-15"
    ws["B6"].number_format = "yyyy-mm-dd"
    # Header drawing: the synthetic logo lives in the sheet header band.
    logo = XLImage(_logo_path())
    logo.width, logo.height = 96, 36
    ws.add_image(logo, "G1")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28


def _build_inputs(wb: Workbook) -> None:
    ws = wb.create_sheet("Inputs")
    ws["A1"] = "Model Inputs"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=ACME_NAVY)
    # Header row 3.
    for col, label in enumerate(("Driver", "Value", "Unit"), start=1):
        c = ws.cell(row=3, column=col, value=label)
        c.style = "AcmeHeader"
    # Input block rows 4..7 (the named INPUTS region, demo values).
    inputs = [
        ("Units sold", 1200, "ea"),
        ("Unit price", 49.5, "USD"),
        ("Discount rate", 0.12, "pct"),
        ("Tax rate", 0.22, "pct"),
    ]
    for i, (driver, val, unit) in enumerate(inputs):
        r = 4 + i
        ws.cell(row=r, column=1, value=driver).style = "AcmeInput"
        vc = ws.cell(row=r, column=2, value=val)
        vc.style = "AcmeInput"
        if unit == "USD":
            vc.style = "AcmeCurrency"
        elif unit == "pct":
            vc.style = "AcmePercent"
        ws.cell(row=r, column=3, value=unit).style = "AcmeInput"
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14


def _build_model(wb: Workbook) -> None:
    ws = wb.create_sheet("Model")
    ws["A1"] = "Revenue Model"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=ACME_NAVY)
    # Header row 3 for the native table.
    headers = ("Line item", "Q1", "Q2", "Q3", "Q4", "FY Total", "% of FY")
    for col, label in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=label).style = "AcmeHeader"
    # Demo body rows 4..7 with cross-quarter SUM + percent-of-total formulas.
    body = [
        ("Gross revenue", 320000, 351000, 372500, 410000),
        ("Discounts", -38400, -42120, -44700, -49200),
        ("Returns", -9600, -10530, -11175, -12300),
        ("Net revenue", None, None, None, None),  # formula row below
    ]
    for i, row in enumerate(body):
        r = 4 + i
        ws.cell(row=r, column=1, value=row[0])
        if row[0] == "Net revenue":
            # Net revenue = sum of the three lines above, per quarter.
            for col in range(2, 6):
                L = get_column_letter(col)
                ws.cell(row=r, column=col, value=f"=SUM({L}4:{L}6)")
        else:
            for col, val in zip(range(2, 6), row[1:]):
                c = ws.cell(row=r, column=col, value=val)
                c.number_format = "#,##0"
        # FY Total (col 6) = SUM of the four quarters in this row.
        ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})").number_format = "#,##0"
        # % of FY (col 7) = this row's FY total / net-revenue FY total (row 7).
        ws.cell(
            row=r, column=7, value=f"=IF($F$7=0,0,F{r}/$F$7)"
        ).number_format = "0.0%"
    # A grand-total SUBTOTAL row 9 (col 6).
    ws.cell(row=9, column=1, value="Subtotal (visible)")
    ws.cell(row=9, column=6, value="=SUBTOTAL(9,F4:F6)").number_format = "#,##0"
    # Native TABLE object over the body (header row 3 .. last data row 7).
    table = Table(displayName="AcmeDataTbl", ref="A3:G7")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
        showFirstColumn=False,
        showLastColumn=False,
    )
    ws.add_table(table)
    # CONDITIONAL FORMATTING: color scale on the quarter grid, a CellIs rule on
    # the % column, and a formula rule that flags negative FY totals.
    ws.conditional_formatting.add(
        "B4:E6",
        ColorScaleRule(
            start_type="min",
            start_color="FFF8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFFFEB84",
            end_type="max",
            end_color="FF63BE7B",
        ),
    )
    ws.conditional_formatting.add(
        "G4:G7",
        CellIsRule(
            operator="greaterThan",
            formula=["0.5"],
            fill=PatternFill("solid", fgColor=ACME_AMBER),
        ),
    )
    ws.conditional_formatting.add(
        "F4:F6",
        FormulaRule(formula=["F4<0"], fill=PatternFill("solid", fgColor="FFFFC7CE")),
    )
    ws.freeze_panes = "B4"
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # A native CHART driven by the model body (bar of quarter columns).
    chart = BarChart()
    chart.title = "Quarterly net revenue"
    chart.type = "col"
    data = Reference(ws, min_col=2, min_row=7, max_col=5, max_row=7)
    cats = Reference(ws, min_col=2, min_row=3, max_col=5, max_row=3)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, "I3")

    # A second chart (line) on the Summary sheet is added there.


def _build_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Executive Summary"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=ACME_NAVY)
    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    ws["A3"].style = "AcmeHeader"
    ws["B3"].style = "AcmeHeader"
    # CROSS-SHEET formulas: Summary pulls from Model and Inputs.
    ws["A4"] = "FY net revenue"
    ws["B4"] = "=Model!F7"
    ws["B4"].number_format = "#,##0"
    ws["A5"] = "Units sold (input)"
    ws["B5"] = "=Inputs!B4"
    ws["A6"] = "Unit price (input)"
    ws["B6"] = "=Inputs!B5"
    ws["B6"].style = "AcmeCurrency"
    ws["A7"] = "Implied gross"
    ws["B7"] = "=Inputs!B4*Inputs!B5"
    ws["B7"].style = "AcmeCurrency"
    ws["A8"] = "Net margin"
    ws["B8"] = "=IF(B7=0,0,B4/B7)"
    ws["B8"].style = "AcmePercent"
    ws["A9"] = "Headline KPI"
    ws["B9"] = "{{headline_kpi}}"  # a single-cell named output slot
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16

    # A LINE chart on the summary referencing the model FY total row.
    model = wb["Model"]
    chart = LineChart()
    chart.title = "FY total by line item"
    data = Reference(model, min_col=6, min_row=3, max_col=6, max_row=7)
    chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, "D3")


def _build_data(wb: Workbook) -> None:
    """A raw transactions sheet (demo data the generator should clear/refill)."""
    ws = wb.create_sheet("Data")
    headers = ("Date", "Region", "Product", "Amount")
    for col, label in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=label).style = "AcmeHeader"
    demo = [
        ("2025-10-01", "North", "Widget", 12500),
        ("2025-10-02", "South", "Gadget", 9800),
        ("2025-10-03", "East", "Widget", 14200),
    ]
    for i, (d, region, prod, amt) in enumerate(demo):
        r = 2 + i
        dc = ws.cell(row=r, column=1, value=d)
        dc.number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=region)
        ws.cell(row=r, column=3, value=prod)
        ac = ws.cell(row=r, column=4, value=amt)
        ac.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    # A total row with a SUM the generator must preserve.
    ws.cell(row=5, column=3, value="Total")
    ws.cell(row=5, column=4, value="=SUM(D2:D4)").number_format = "#,##0"
    ws.freeze_panes = "A2"
    for col, w in zip("ABCD", (14, 12, 14, 16)):
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------------------
# Named ranges (workbook scope) - the author's OWN vocabulary, surfaced as
# geometry by the extractor (never matched on as code literals).
# ---------------------------------------------------------------------------
def _add_named_ranges(wb: Workbook) -> None:
    defns = {
        # Single-cell title anchor under a merged header band.
        "report_title": "'Cover'!$A$1",
        "report_subtitle": "'Cover'!$A$2",
        "client_name": "'Cover'!$B$4",
        "period": "'Cover'!$B$5",
        # Multi-cell INPUTS block (sample-data candidate, in the frozen band edge).
        "inputs_block": "'Inputs'!$A$4:$C$7",
        # Multi-cell DATA region = body of the native table.
        "model_body": "'Model'!$A$4:$G$6",
        # A cross-sheet formula OUTPUT cell (single-cell named output slot).
        "headline_kpi": "'Summary'!$B$9",
        # The raw demo-data block.
        "data_block": "'Data'!$A$2:$D$4",
    }
    for name in sorted(defns):
        wb.defined_names.add(DefinedName(name, attr_text=defns[name]))


_LOGO_CACHE: Path | None = None


def _logo_path() -> Path:
    """Materialize the synthetic logo to a temp file openpyxl can embed."""
    global _LOGO_CACHE
    if _LOGO_CACHE is None:
        import tempfile

        tmp = Path(tempfile.gettempdir()) / "acme_complex_logo.png"
        tmp.write_bytes(_synthetic_logo_png())
        _LOGO_CACHE = tmp
    return _LOGO_CACHE


def build(out: Path = OUT) -> Path:
    wb = Workbook()
    # Drop the default sheet; we author named sheets in a deliberate order.
    wb.remove(wb.active)
    _register_named_styles(wb)
    _build_cover(wb)
    _build_inputs(wb)
    _build_model(wb)
    _build_summary(wb)
    _build_data(wb)
    _add_named_ranges(wb)
    # Workbook-level: request a full recalc so authored formulas evaluate on open.
    wb.calculation.fullCalcOnLoad = True
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


if __name__ == "__main__":
    path = build()
    print(f"built {path}")
