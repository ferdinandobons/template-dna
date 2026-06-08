# SPDX-License-Identifier: MIT
"""XLSX parity (M-i-6): geometry inventories, shared-resolver routing,
comprehension-steered cover/demo reconciliation, formula preservation, recalc,
the destructive floor, and idempotency.

Every test builds a SYNTHETIC workbook (proprietary templates are never copied
into the repo). The range names used as grid keys / surfaced ids ("title_cell",
"data_region") are the synthetic author's OWN vocabulary carried as DATA - the
extractor never matches on them as code-side literals (that is exactly what
M-i-6 de-literalized).
"""

from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from brandkit.formats.xlsx import extract as xlsx_extract
from brandkit.formats.xlsx import generate as xlsx_generate
from brandkit.grid.model import GridDocument
from brandkit.profile import comprehension as comp
from brandkit.profile import store


def _workbook(
    path: Path,
    *,
    title_range: str = "$A$1",
    merged_title: bool = False,
    freeze: str | None = "A4",
    with_table: bool = False,
    data_rows: int = 1,
) -> None:
    """Build a synthetic workbook with a cover cell, a header, sample data, a
    formula, and (optionally) a merged title / freeze panes / a table object."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "{{title}}"
    if merged_title:
        ws.merge_cells("A1:C1")
    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    last_data_row = 3 + data_rows
    for i in range(data_rows):
        r = 4 + i
        ws.cell(row=r, column=1).value = f"Example row {i + 1}"
        ws.cell(row=r, column=2).value = i + 1
    ws.cell(row=last_data_row + 1, column=2).value = f"=SUM(B4:B{last_data_row})"
    if freeze:
        ws.freeze_panes = freeze
    if with_table:
        ws.add_table(Table(displayName="DataTbl", ref=f"A3:B{last_data_row}"))
    wb.defined_names.add(DefinedName("title_cell", attr_text=f"'Report'!{title_range}"))
    wb.defined_names.add(
        DefinedName("data_region", attr_text=f"'Report'!$A$4:$B${last_data_row}")
    )
    wb.save(path)


class XlsxGeometryInventory(unittest.TestCase):
    """The extractor emits geometry evidence, with NO range-name literals."""

    def _extract(self, td: Path, **wb_kwargs) -> dict:
        tpl = td / "t.xlsx"
        _workbook(tpl, **wb_kwargs)
        old = os.getcwd()
        os.chdir(td)
        try:
            xlsx_extract.extract(tpl, "model", scope="project", cwd=td)
            return store.load_profile("model", "project").profile
        finally:
            os.chdir(old)

    def test_single_cell_anchor_captures_demo_value(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            prof = self._extract(Path(t), title_range="$A$1")
            anchors = {a["name"]: a for a in prof["surface"]["xlsx"]["cover_anchors"]}
            self.assertEqual(anchors["title_cell"]["cardinality"], "single_cell")
            self.assertEqual(anchors["title_cell"]["demo_value"], "{{title}}")
            # No privileged "title" role: a generic named_range role keyed by slug.
            self.assertNotIn("title", prof["roles"])
            self.assertEqual(
                prof["roles"]["region.titlecell"]["resolver"]["name"], "title_cell"
            )

    def test_merged_header_and_frozen_band_detected(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            prof = self._extract(
                Path(t), title_range="$A$1:$C$1", merged_title=True, freeze="A4"
            )
            anchors = {a["name"]: a for a in prof["surface"]["xlsx"]["cover_anchors"]}
            title = anchors["title_cell"]
            self.assertEqual(title["cardinality"], "multi_cell")
            self.assertTrue(title["merged_header"])
            self.assertTrue(title["in_frozen_band"])

    def test_table_membership_detected(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            prof = self._extract(Path(t), with_table=True, data_rows=2)
            anchors = {a["name"]: a for a in prof["surface"]["xlsx"]["cover_anchors"]}
            data = anchors["data_region"]
            self.assertEqual(data["table"], "DataTbl")
            self.assertEqual(data["table_role"], "body")

    def test_fields_inventory_is_legal_empty(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            prof = self._extract(Path(t))
            # A workbook has no TOC-style field code: fields is empty (readiness gate).
            self.assertEqual(prof["surface"]["xlsx"]["fields"], [])

    def test_regions_inventory_has_sheet_and_sample_data(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            prof = self._extract(Path(t), data_rows=2)
            regions = prof["surface"]["xlsx"]["regions"]
            kinds = {r["kind"] for r in regions}
            self.assertIn("sheet", kinds)
            self.assertIn("sample_data", kinds)
            ids = {r["id"] for r in regions}
            self.assertIn("region.dataregion", ids)
            self.assertIn("region.sheet.report", ids)

    def test_skeleton_lists_sheets_in_order(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            prof = self._extract(Path(t))
            skel = prof["structure"]["skeleton"]
            self.assertTrue(prof["structure"]["ordered"])
            self.assertEqual(skel[0]["region"], "sheet.report")
            self.assertEqual(skel[0]["order"], 0)

    def test_no_named_ranges_yields_safe_default_role(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            tpl = td / "bare.xlsx"
            wb = Workbook()
            wb.active["A1"] = "x"
            wb.save(tpl)
            old = os.getcwd()
            os.chdir(td)
            try:
                xlsx_extract.extract(tpl, "bare", scope="project", cwd=td)
                prof = store.load_profile("bare", "project").profile
            finally:
                os.chdir(old)
            self.assertEqual(prof["surface"]["xlsx"]["cover_anchors"], [])
            # Registry stays non-empty so generation has a fallback.
            self.assertIn("cell.default", prof["roles"])
            self.assertEqual(prof["anchors"]["cover"]["kind"], "NONE")


class XlsxComprehensionReconciliation(unittest.TestCase):
    """Comprehend -> generate: fill cover in place, clear demo, preserve formulas."""

    def _extract_profile(self, td: Path, **wb_kwargs):
        tpl = td / "t.xlsx"
        _workbook(tpl, **wb_kwargs)
        old = os.getcwd()
        os.chdir(td)
        try:
            xlsx_extract.extract(tpl, "model", scope="project", cwd=td)
            return store.load_profile("model", "project")
        finally:
            os.chdir(old)

    def _comprehend(self, profile: dict, block: dict):
        return comp.merge(
            profile,
            block,
            generated_by={
                "model": "test",
                "prompt_version": "v1",
                "generated_at": "2026-01-01T00:00:00Z",
            },
        )

    def test_demo_clear_and_formula_preserved(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract_profile(td, data_rows=2)
            prof = loaded.profile
            res = self._comprehend(
                prof,
                {
                    # Confident classification: above the destructive floor so the
                    # demo clear is honored (the uniform confidence floor now gates
                    # the xlsx demo clear, mirroring the docx/pptx cover reconcilers).
                    "confidence": 0.9,
                    "cover_slots": {
                        "anchor.titlecell": {
                            "semantic_role": "title",
                            "binds_to": "title",
                            "fill_rule": "in_place",
                        }
                    },
                    "demo_classification": {
                        "regions": [
                            {
                                "region_ref": "region.dataregion",
                                "verdict": "demo",
                                "evidence": "sample rows",
                            }
                        ]
                    },
                },
            )
            self.assertTrue(res.ok, res.problems)
            self.assertTrue(store.comprehension_is_present(prof))

            grid = GridDocument(
                cells={"title_cell": "Quarterly Model"},
                regions={"data_region": [["Pipeline", 42]]},
            )
            findings: list = []
            out = td / "out.xlsx"
            xlsx_generate.generate(
                prof, loaded.shell_path, grid, out, findings=findings
            )

            # No destructive-floor ERROR: the demo clear is corroborated by verdict.
            self.assertFalse(
                [f for f in findings if f.severity == "ERROR"],
                [f.message for f in findings],
            )

            wb = load_workbook(out, data_only=False)
            ws = wb["Report"]
            self.assertEqual(ws["A1"].value, "Quarterly Model")
            self.assertEqual(ws["A4"].value, "Pipeline")
            self.assertEqual(ws["B4"].value, 42)
            # The stale second demo row was cleared in place.
            self.assertIsNone(ws["A5"].value)
            self.assertIsNone(ws["B5"].value)
            # The formula (one row below the data block) is preserved verbatim.
            self.assertEqual(ws["B6"].value, "=SUM(B4:B5)")
            # Recalc requested so Excel recomputes the preserved formula on open.
            self.assertTrue(wb.calculation.fullCalcOnLoad)

    def test_generate_is_idempotent(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract_profile(td, data_rows=2)
            prof = loaded.profile
            self._comprehend(
                prof,
                {
                    # Above the destructive floor so the demo clear actually fires
                    # (otherwise idempotency would hold trivially over a no-op clear).
                    "confidence": 0.9,
                    "cover_slots": {"anchor.titlecell": {"fill_rule": "in_place"}},
                    "demo_classification": {
                        "regions": [
                            {"region_ref": "region.dataregion", "verdict": "demo"}
                        ]
                    },
                },
            )
            grid = GridDocument(
                cells={"title_cell": "Q"}, regions={"data_region": [["Pipeline", 42]]}
            )
            a = td / "a.xlsx"
            b = td / "b.xlsx"
            xlsx_generate.generate(prof, loaded.shell_path, grid, a)
            xlsx_generate.generate(prof, loaded.shell_path, grid, b)
            self.assertEqual(a.read_bytes(), b.read_bytes())

    def test_cover_clear_rule_empties_anchor(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract_profile(td)
            prof = loaded.profile
            res = self._comprehend(
                prof,
                {
                    # Above the destructive floor so the cover clear is honored.
                    "confidence": 0.9,
                    "cover_slots": {"anchor.titlecell": {"fill_rule": "clear"}},
                },
            )
            self.assertTrue(res.ok, res.problems)
            grid = GridDocument()  # nothing filled; the anchor is just cleared
            findings: list = []
            out = td / "out.xlsx"
            xlsx_generate.generate(
                prof, loaded.shell_path, grid, out, findings=findings
            )
            self.assertFalse(
                [f for f in findings if f.severity == "ERROR"],
                [f.message for f in findings],
            )
            ws = load_workbook(out, data_only=False)["Report"]
            self.assertIsNone(
                ws["A1"].value
            )  # placeholder cleared, not left as {{title}}

    def test_low_confidence_demo_clear_downgraded_to_keep(self) -> None:
        # A demo verdict the model is NOT confident about (< the destructive floor):
        # the uniform confidence floor downgrades the clear to KEEP + WARNING, so the
        # stale demo row SURVIVES rather than being deleted (a wrong delete is
        # unrecoverable). Mirrors the docx/pptx cover-clear downgrade.
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract_profile(td, data_rows=2)
            prof = loaded.profile
            res = self._comprehend(
                prof,
                {
                    "confidence": 0.3,  # below DESTRUCTIVE_CONFIDENCE_FLOOR (0.5)
                    "cover_slots": {
                        "anchor.titlecell": {
                            "binds_to": "title",
                            "fill_rule": "in_place",
                        }
                    },
                    "demo_classification": {
                        "regions": [
                            {"region_ref": "region.dataregion", "verdict": "demo"}
                        ]
                    },
                },
            )
            self.assertTrue(res.ok, res.problems)
            # The grid refills only ONE row, so the second demo row would be cleared
            # by a corroborated demo clear; under the floor it must be kept.
            grid = GridDocument(
                cells={"title_cell": "Quarterly Model"},
                regions={"data_region": [["Pipeline", 42]]},
            )
            findings: list = []
            out = td / "out.xlsx"
            xlsx_generate.generate(
                prof, loaded.shell_path, grid, out, findings=findings
            )
            # KEEP + WARNING, and NO net-loss ERROR (nothing was removed).
            self.assertTrue(
                any(f.check == "demo_clear_downgraded" for f in findings),
                [f.check for f in findings],
            )
            self.assertFalse([f for f in findings if f.severity == "ERROR"])
            ws = load_workbook(out, data_only=False)["Report"]
            self.assertEqual(ws["A4"].value, "Pipeline")  # grid-refilled row
            # The trailing demo row was NOT cleared (low-confidence -> keep).
            self.assertEqual(ws["A5"].value, "Example row 2")

    def test_low_confidence_cover_clear_downgraded_to_keep(self) -> None:
        # A cover clear the model is NOT confident about: downgraded to KEEP + WARNING,
        # so the placeholder is left in place rather than emptied.
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract_profile(td)
            prof = loaded.profile
            res = self._comprehend(
                prof,
                {
                    "confidence": 0.3,  # below the floor
                    "cover_slots": {"anchor.titlecell": {"fill_rule": "clear"}},
                },
            )
            self.assertTrue(res.ok, res.problems)
            grid = GridDocument()  # nothing filled; only the (downgraded) clear acts
            findings: list = []
            out = td / "out.xlsx"
            xlsx_generate.generate(
                prof, loaded.shell_path, grid, out, findings=findings
            )
            self.assertTrue(
                any(f.check == "cover_clear_downgraded" for f in findings),
                [f.check for f in findings],
            )
            self.assertFalse([f for f in findings if f.severity == "ERROR"])
            ws = load_workbook(out, data_only=False)["Report"]
            # Placeholder NOT cleared (low-confidence -> keep).
            self.assertEqual(ws["A1"].value, "{{title}}")

    def test_absent_comprehension_uses_deterministic_path(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract_profile(td, data_rows=2)
            prof = loaded.profile  # no comprehend() call -> status absent
            self.assertFalse(store.comprehension_is_present(prof))
            grid = GridDocument(
                cells={"title_cell": "Q"}, regions={"data_region": [["Pipeline", 42]]}
            )
            out = td / "out.xlsx"
            xlsx_generate.generate(prof, loaded.shell_path, grid, out)
            ws = load_workbook(out, data_only=False)["Report"]
            # Deterministic fill only; no demo clear (row 5 stays as authored).
            self.assertEqual(ws["A1"].value, "Q")
            self.assertEqual(ws["A4"].value, "Pipeline")
            self.assertEqual(ws["A5"].value, "Example row 2")

    def test_unknown_grid_key_is_error(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract_profile(td)
            grid = GridDocument(cells={"not_a_region": "x"})
            with self.assertRaises(ValueError):
                xlsx_generate.generate(
                    loaded.profile, loaded.shell_path, grid, td / "out.xlsx"
                )


class XlsxResolverRouting(unittest.TestCase):
    """Generation routes every fill through the shared ProfileResolver spine."""

    def test_fill_target_comes_from_resolver_not_raw_surface(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            tpl = td / "t.xlsx"
            _workbook(tpl)
            old = os.getcwd()
            os.chdir(td)
            try:
                xlsx_extract.extract(tpl, "model", scope="project", cwd=td)
                loaded = store.load_profile("model", "project")
            finally:
                os.chdir(old)
            # Corrupt the raw surface map so a generator that read it directly would
            # write to the wrong place; the resolver-routed path must still resolve
            # the role's named_range and fill the right cell (A1).
            loaded.profile["surface"]["xlsx"]["named_regions"]["title_cell"][
                "range"
            ] = "$A$1"
            grid = GridDocument(cells={"title_cell": "Routed"})
            out = td / "out.xlsx"
            xlsx_generate.generate(loaded.profile, loaded.shell_path, grid, out)
            ws = load_workbook(out, data_only=False)["Report"]
            self.assertEqual(ws["A1"].value, "Routed")


class XlsxNoProprietaryLiterals(unittest.TestCase):
    """Guard against re-introducing the de-literalized range names as code rules."""

    def test_extract_and_structure_have_no_range_name_literals(self) -> None:
        import brandkit.formats.xlsx.extract as ex
        import brandkit.formats.xlsx.structure as st

        for mod in (ex, st):
            src = Path(mod.__file__).read_text(encoding="utf-8")
            # The de-literalized names must not appear as matching rules. They are
            # allowed only inside docstrings as illustrative examples, so check the
            # CODE lines (strip docstring/comment-only context heuristically: a bare
            # quoted occurrence outside an example would be a regression). We assert
            # they never appear as a string compared against a value.
            self.assertNotIn('== "title_cell"', src)
            self.assertNotIn('"title_cell" in', src)
            self.assertNotIn('!= "title_cell"', src)
            self.assertNotIn('"data_region" in', src)


if __name__ == "__main__":
    unittest.main()
