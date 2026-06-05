# SPDX-License-Identifier: MIT
"""XLSX fidelity regression tests against the COMPLEX synthetic fixture.

These assert the X-group strengthening fixes on the realistic
``tests/fixtures/complex/acme_complex.xlsx`` workbook (multi-sheet, cross-sheet
formulas, native table + charts + CF + image, named cell styles, demo data):

  * X1   - refilling a named region that straddles formula cells preserves the
           formulas verbatim (the headline data-corruption bug) AND the QA
           formula-preservation check would have caught it.
  * X2/X3/X5 - the enriched extraction inventories (number formats, named cell
           styles + cell_style roles, table styles, charts, CF, images) are real,
           not hardcoded ``[]``.
  * CC-1 - ``comprehend-input`` yields a non-empty excerpt + ``facts.styles`` for
           the workbook (previously empty for every xlsx).
  * X4   - a filled cover anchor keeps its brand named cell style.
  * X7   - two identical generations are byte-identical (pinned save timestamp).

The fixture is 100% synthetic ("Acme"); no proprietary template is copied in.
The range / style names used as ids are the synthetic author's OWN vocabulary
carried as data - never matched as code-side literals.
"""
from __future__ import annotations

import os
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "scripts"))

from openpyxl import load_workbook

from brandkit.formats.xlsx import extract as xlsx_extract
from brandkit.formats.xlsx import generate as xlsx_generate
from brandkit.grid.model import GridDocument
from brandkit.profile import comprehension as comp
from brandkit.profile import schema
from brandkit.profile import store
from brandkit.qa import checks_deterministic as checks

FIXTURE = Path(__file__).resolve().parents[1] / "tests" / "fixtures" / "complex" / "acme_complex.xlsx"


def _formula_map(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    wb = load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        for cell in ws._cells.values():
            if isinstance(cell.value, str) and cell.value.startswith("="):
                out[f"{ws.title}!{cell.coordinate}"] = cell.value
    return out


class _ComplexBase(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        if not FIXTURE.exists():
            raise unittest.SkipTest(f"missing complex fixture {FIXTURE}")

    def _extract(self, td: Path):
        old = os.getcwd()
        os.chdir(td)
        try:
            xlsx_extract.extract(FIXTURE, "acme", scope="project", cwd=td)
            return store.load_profile("acme", "project")
        finally:
            os.chdir(old)


class XlsxFormulaPreservation(_ComplexBase):
    """X1: a wide region fill over formula columns never erases the formulas."""

    def test_region_fill_over_formula_columns_preserves_formulas(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td)
            prof = loaded.profile
            shell_formulas = prof["artifact_catalog"]["formulas"]
            self.assertEqual(len(shell_formulas), 19)
            # model_body = Model!$A$4:$G$6 overlaps the F4:F6 (=SUM) and G4:G6 (=IF)
            # formula columns. A FULL-WIDTH refill used to blank them (19 -> 13).
            grid = GridDocument(
                regions={
                    "model_body": [
                        ["Gross revenue", 1, 2, 3, 4, "", ""],
                        ["Discounts", -1, -2, -3, -4, "", ""],
                        ["Returns", -1, -1, -1, -1, None, None],
                    ]
                }
            )
            out = td / "out.xlsx"
            xlsx_generate.generate(prof, loaded.shell_path, grid, out)

            out_formulas = _formula_map(out)
            # No formula lost or mutated.
            self.assertEqual(set(shell_formulas), set(out_formulas))
            for addr, f in shell_formulas.items():
                self.assertEqual(out_formulas[addr], f, addr)
            # The literal cells DID get written (the fill still works).
            wb = load_workbook(out, data_only=False)
            self.assertEqual(wb["Model"]["A4"].value, "Gross revenue")
            self.assertEqual(wb["Model"]["B4"].value, 1)
            # The formula columns are untouched.
            self.assertEqual(wb["Model"]["F4"].value, "=SUM(B4:E4)")
            self.assertEqual(wb["Model"]["G4"].value, "=IF($F$7=0,0,F4/$F$7)")

    def test_qa_formula_preservation_passes_on_guarded_output(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td)
            prof = loaded.profile
            grid = GridDocument(
                regions={"model_body": [["X", 1, 2, 3, 4, "", ""]]}
            )
            out = td / "out.xlsx"
            xlsx_generate.generate(prof, loaded.shell_path, grid, out)
            findings = checks.check_formula_preservation(loaded.shell_path, out, prof)
            errors = [f for f in findings if f.severity == "ERROR"]
            self.assertEqual(errors, [], [f.message for f in errors])

    def test_qa_formula_preservation_catches_an_erased_formula(self) -> None:
        # A deliberately corrupted output (a formula blanked) MUST be flagged, so
        # the gate - not Excel - is the detector going forward.
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td)
            prof = loaded.profile
            out = td / "out.xlsx"
            wb = load_workbook(loaded.shell_path, data_only=False)
            wb["Model"]["F4"].value = None  # erase a SUM formula
            wb.save(out)
            findings = checks.check_formula_preservation(loaded.shell_path, out, prof)
            errors = [f for f in findings if f.severity == "ERROR"]
            self.assertTrue(errors, "erased formula must be flagged")
            self.assertIn("Model!F4", " ".join(f.message for f in errors))


class XlsxEnrichedInventories(_ComplexBase):
    """X2/X3/X5: extraction surfaces real component inventories (not []) + roles."""

    def test_number_formats_surfaced(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            prof = self._extract(Path(t)).profile
            masks = {n["format"] for n in prof["surface"]["xlsx"]["number_formats"]}
            self.assertIn("0.0%", masks)
            self.assertIn("yyyy-mm-dd", masks)
            self.assertIn("#,##0", masks)
            # Each entry carries cardinality + a sample address.
            for entry in prof["surface"]["xlsx"]["number_formats"]:
                self.assertGreaterEqual(entry["count"], 1)
                self.assertIn("!", entry["sample"])

    def test_named_cell_styles_and_roles(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            prof = self._extract(Path(t)).profile
            names = {s["name"] for s in prof["surface"]["xlsx"]["named_styles"]}
            self.assertIn("AcmeTitle", names)
            self.assertIn("AcmeCurrency", names)
            # A brand style is promoted to a cell_style role with a verbatim target.
            roles = prof["roles"]
            self.assertIn("cell.style.acmetitle", roles)
            self.assertEqual(
                roles["cell.style.acmetitle"]["resolver"]["type"], "cell_style"
            )
            self.assertEqual(
                roles["cell.style.acmetitle"]["resolver"]["style_name"], "AcmeTitle"
            )
            # The builtin 'Normal' is NOT nominated.
            self.assertNotIn("cell.style.normal", roles)
            # Every emitted role id must satisfy the shared role-id grammar
            # (dotted lowercase alphanumeric, NO underscores) so the schema
            # validator / QA gate accept the profile - regression for the
            # "cell_style.*" (underscore) family that schema.validate rejected.
            self.assertEqual(
                [e for e in schema.validate(prof) if "invalid role id" in e], []
            )

    def test_table_charts_cf_image_inventories(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            prof = self._extract(Path(t)).profile
            sx = prof["surface"]["xlsx"]
            tables = {tbl["name"]: tbl for tbl in sx["table_styles"]}
            self.assertIn("AcmeDataTbl", tables)
            self.assertEqual(tables["AcmeDataTbl"]["style"], "TableStyleMedium2")
            chart_types = {c["type"] for c in sx["charts"]}
            self.assertEqual(chart_types, {"BarChart", "LineChart"})
            self.assertTrue(sx["images"])  # the header logo
            cf_sqrefs = {c["sqref"] for c in sx["conditional_formatting"]}
            self.assertIn("B4:E6", cf_sqrefs)

    def test_catalog_has_component_baselines(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            cat = self._extract(Path(t)).profile["artifact_catalog"]
            self.assertTrue(cat["charts"])
            self.assertTrue(cat["tables"])
            self.assertTrue(cat["images"])


class XlsxComprehendInputExcerpt(_ComplexBase):
    """CC-1: the model bundle yields a real excerpt + facts.styles for xlsx."""

    def test_excerpt_and_styles_non_empty(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            prof = self._extract(Path(t)).profile
            bundle = comp.comprehend_input_bundle(prof)
            self.assertTrue(bundle["excerpt"], "xlsx excerpt must not be empty")
            # Real cell text (not just addresses) is present.
            self.assertIn("{{report_title}}", bundle["excerpt"])
            self.assertTrue(bundle["facts"]["styles"])
            self.assertIn("AcmeTitle", bundle["facts"]["styles"]["named_styles"])


class XlsxCoverStyleAndIdempotency(_ComplexBase):
    """X4 cover-cell style re-assertion + X7 byte idempotency."""

    def test_filled_cover_keeps_brand_style(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td)
            grid = GridDocument(cells={"report_title": "New Report Title"})
            out = td / "out.xlsx"
            xlsx_generate.generate(loaded.profile, loaded.shell_path, grid, out)
            ws = load_workbook(out, data_only=False)["Cover"]
            self.assertEqual(ws["A1"].value, "New Report Title")
            self.assertEqual(ws["A1"].style, "AcmeTitle")

    def test_generation_is_byte_idempotent(self) -> None:
        with tempfile.TemporaryDirectory() as t:
            td = Path(t)
            loaded = self._extract(td)
            grid = GridDocument(
                cells={"report_title": "T"},
                regions={"model_body": [["x", 1, 2, 3, 4, "", ""]]},
            )
            a = td / "a.xlsx"
            b = td / "b.xlsx"
            xlsx_generate.generate(loaded.profile, loaded.shell_path, grid, a)
            xlsx_generate.generate(loaded.profile, loaded.shell_path, grid, b)
            self.assertEqual(a.read_bytes(), b.read_bytes())


if __name__ == "__main__":
    unittest.main()
