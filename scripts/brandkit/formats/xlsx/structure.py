# SPDX-License-Identifier: MIT
"""XLSX structure helpers - the peer of ``formats/docx/structure.py``.

The workbook's *structure* (named regions, sheet geometry, sample-data blocks) is
surfaced here as the format-uniform inventories the comprehension model reasons
over and the validator binds to (plan §4):

  - :func:`inventory_cover_anchors` - every named region's **geometry**, evidence
    only: cardinality (single-cell vs multi-cell), whether it straddles a merged
    range (a merged header row), whether it sits in the frozen header band, and
    whether it is the anchor row of a table object. The captured cell text is the
    region's ``demo_value`` (a verbatim placeholder the generator can match the
    residual-text guard against). This **replaces** the ``"title_cell"`` literal:
    purpose is the model's job (``cover_slots``), never a hardcoded name.
  - :func:`inventory_regions` - the sheets and the sample-data ranges. Each named
    multi-cell region is a sample-data candidate; plus each sheet's contiguous
    body block under a (frozen / styled) header is surfaced. These back
    ``demo_classification`` / ``conventions.sections`` refs.
  - :func:`inventory_fields` - xlsx has no TOC-style field codes; it is empty (a
    legal, format-uniform shape: the ``fields`` inventory simply has nothing for
    the model to bind an ``index_ref`` to, so an xlsx ``conventions.indexes`` ref
    is fail-closed at QA time, per the readiness gate).
  - :func:`detect_skeleton` - the ordered sheet skeleton (the ``structure``
    payload), peer of the docx cover/toc/body skeleton.

Detection is grounded in **geometry evidence**, never in brand-specific range
names. A region id is derived from the named range's OWN name (the template
author's data carried into the profile as an id, exactly as docx surfaces an
opaque ``\\c`` seq_id) - never matched against a code-side word list. So this
works for any workbook in any language.
"""

from __future__ import annotations

import re
from typing import Any, Optional

from openpyxl.utils.cell import range_boundaries

from brandkit.common.text import slugify


# ---------------------------------------------------------------------------
# Stable id derivation (the named range's own name -> a syntactically valid id)
# ---------------------------------------------------------------------------
def region_id_for_name(name: str) -> str:
    """Return the stable inventory id for a named range.

    The id is derived from the range's OWN name via :func:`slugify` (hyphens
    stripped so the result is a valid dotted role/region token), prefixed with
    ``region.`` so it shares the docx region-ref namespace
    (``region.cover`` / ``region.field.18`` / ``region.<slug>``). The range name
    is the template author's data, not a code literal - we never pattern-match on
    it, we only carry it as an id the model binds to.
    """
    return f"region.{slugify(name).replace('-', '')}"


def anchor_id_for_name(name: str) -> str:
    """Return the stable cover-anchor id for a named range.

    Mirrors :func:`region_id_for_name` in the ``anchor.`` namespace so a
    ``cover_slots`` key (an ``anchor_ref``) and a ``demo_classification`` /
    ``sections`` key (a ``region_ref``) over the SAME named range stay distinct
    inventories while both deriving deterministically from the range's own name.
    """
    return f"anchor.{slugify(name).replace('-', '')}"


# ---------------------------------------------------------------------------
# Named-region geometry (the deterministic evidence; no literals)
# ---------------------------------------------------------------------------
def _first_destination(defined_name) -> Optional[tuple[str, str]]:
    """Return the first ``(sheet, coord)`` destination of a defined name, or None."""
    try:
        destinations = list(defined_name.destinations)
    except Exception:
        return None
    if not destinations:
        return None
    sheet, coord = destinations[0]
    return sheet, coord


def _merged_ranges(ws) -> list:
    """Return the worksheet's merged-cell range objects (empty when none)."""
    try:
        return list(ws.merged_cells.ranges)
    except Exception:
        return []


def _freeze_bounds(ws) -> Optional[tuple[int, int]]:
    """Return ``(freeze_row, freeze_col)`` 1-based for the freeze split, or None.

    ``ws.freeze_panes`` is the top-left UNFROZEN cell; everything strictly above
    its row is the frozen header band and everything strictly left of its column
    is the frozen index band. Returns the split cell's ``(row, col)``.
    """
    fp = ws.freeze_panes
    if not fp:
        return None
    try:
        col, row, _, _ = range_boundaries(str(fp))
    except Exception:
        return None
    return row, col


def _table_anchor_rows(ws) -> dict[str, tuple[int, int, int, int]]:
    """Return ``{table_name: (min_row, max_row, min_col, max_col)}`` for tables."""
    out: dict[str, tuple[int, int, int, int]] = {}
    try:
        items = list(ws.tables.items())
    except Exception:
        return out
    for tname, ref in items:
        coord = ref.ref if hasattr(ref, "ref") else ref
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(coord))
        except Exception:
            continue
        out[str(tname)] = (min_row, max_row, min_col, max_col)
    return out


def _cell_text(ws, row: int, col: int) -> Optional[str]:
    """Return the cell's string value (the captured demo text), or None.

    Only a *string* value is a placeholder/demo text the residual-text guard can
    match; a number/formula/None is not surfaced as ``demo_value``.
    """
    try:
        value = ws.cell(row=row, column=col).value
    except Exception:
        return None
    return value if isinstance(value, str) else None


def _region_geometry(wb, sheet: str, coord: str) -> Optional[dict]:
    """Compute the geometry evidence for one named range, or None if unresolvable."""
    if sheet not in wb.sheetnames:
        return None
    ws = wb[sheet]
    try:
        min_col, min_row, max_col, max_row = range_boundaries(coord)
    except Exception:
        return None
    n_rows = max_row - min_row + 1
    n_cols = max_col - min_col + 1
    single_cell = n_rows == 1 and n_cols == 1

    # Merged-header evidence: does the region's first row intersect a merged range?
    merged_header = False
    for mr in _merged_ranges(ws):
        if mr.min_row <= min_row <= mr.max_row and not (
            mr.max_col < min_col or mr.min_col > max_col
        ):
            merged_header = True
            break

    # Frozen-band membership: does the region sit (partly) in the frozen header
    # rows or frozen index columns?
    in_frozen = False
    fb = _freeze_bounds(ws)
    if fb is not None:
        frow, fcol = fb
        if (frow > 1 and min_row < frow) or (fcol > 1 and min_col < fcol):
            in_frozen = True

    # Table membership: is the region the header (top) row, or body, of a table?
    table_name = None
    table_role = None
    for tname, (tmin_r, tmax_r, tmin_c, tmax_c) in _table_anchor_rows(ws).items():
        overlaps = not (
            max_row < tmin_r or min_row > tmax_r or max_col < tmin_c or min_col > tmax_c
        )
        if not overlaps:
            continue
        table_name = tname
        table_role = "header" if min_row <= tmin_r <= max_row else "body"
        break

    demo_value = _cell_text(ws, min_row, min_col) if single_cell else None

    geo: dict[str, Any] = {
        "sheet": sheet,
        "range": coord,
        "cardinality": "single_cell" if single_cell else "multi_cell",
        "rows": n_rows,
        "cols": n_cols,
        "merged_header": merged_header,
        "in_frozen_band": in_frozen,
    }
    if table_name is not None:
        geo["table"] = table_name
        geo["table_role"] = table_role
    if demo_value is not None:
        geo["demo_value"] = demo_value
    return geo


def _iter_named_regions(wb) -> list[tuple[str, str, str]]:
    """Yield ``(name, sheet, coord)`` for every workbook-level named range, in
    a stable (sorted-by-name) order so the surfaced inventories are deterministic."""
    out: list[tuple[str, str, str]] = []
    try:
        items = wb.defined_names.items()
    except AttributeError:
        # Older openpyxl exposes defined_names as a list-like of DefinedName.
        items = [(dn.name, dn) for dn in wb.defined_names]
    for name, defined_name in items:
        dest = _first_destination(defined_name)
        if dest is None:
            continue
        sheet, coord = dest
        out.append((str(name), sheet, coord))
    out.sort(key=lambda t: t[0])
    return out


def named_regions_map(wb) -> dict[str, dict]:
    """Return ``{name: {"sheet", "range"}}`` for every named range (back-compat).

    This is the literal-keyed map the deterministic generator fills from when
    comprehension is absent: the grid's ``cells``/``regions`` keys are the named
    range names verbatim (the author's vocabulary), so the deterministic path
    needs no code-side literal - it just looks the author's own name up here.
    """
    return {
        name: {"sheet": sheet, "range": coord}
        for name, sheet, coord in _iter_named_regions(wb)
    }


# ---------------------------------------------------------------------------
# Format-uniform inventories (peers of docx structure.inventory_*)
# ---------------------------------------------------------------------------
def inventory_cover_anchors(wb) -> list[dict]:
    """Surface every named region's geometry as a cover-anchor inventory entry.

    One entry per named range::

        {"id": "anchor.titlecell", "name": "title_cell", "sheet": "Report",
         "range": "$A$1", "cardinality": "single_cell", "rows": 1, "cols": 1,
         "merged_header": false, "in_frozen_band": false,
         "demo_value": "{{title}}"}

    The model binds a ``cover_slots`` key to one of these ids and assigns its
    purpose; the validator checks membership. Evidence only - the id derives from
    the range's OWN name, never from a code word-list, so it is brand- and
    language-agnostic.
    """
    out: list[dict] = []
    for name, sheet, coord in _iter_named_regions(wb):
        geo = _region_geometry(wb, sheet, coord)
        if geo is None:
            continue
        entry = {"id": anchor_id_for_name(name), "name": name}
        entry.update(geo)
        out.append(entry)
    return out


def inventory_fields(wb) -> list[dict]:
    """Surface the field/index inventory.

    A workbook has no TOC-style complex field, so this is ALWAYS empty - a legal,
    format-uniform shape (the docx peer returns the TOC/index fields; xlsx simply
    has none). An xlsx ``conventions.indexes`` ref therefore has nothing to bind
    to and is fail-closed at QA time (the readiness gate: cover/index stay
    effectively out of scope for xlsx; only ``demo_classification`` is grounded).
    """
    return []


def inventory_regions(wb) -> list[dict]:
    """Surface the workbook's region inventory (sheets + sample-data ranges).

    Combines:
      - one ``region.sheet.<slug>`` per worksheet (a structural region a
        ``conventions.sections`` ref can target), and
      - one ``region.<slug>`` per multi-cell named range (a sample-data candidate
        a ``demo_classification`` ref can target as ``verdict=demo``).

    Each entry is ``{"id": <region_ref>, "kind": <open advisory token>, ...}``.
    Deterministic and recomputable at generate time; the ids never encode brand
    words, only the author's own sheet/range names carried as data.
    """
    out: list[dict] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        fb = _freeze_bounds(ws)
        out.append(
            {
                "id": f"region.sheet.{slugify(sheet).replace('-', '')}",
                "kind": "sheet",
                "name": sheet,
                "sheet": sheet,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "has_frozen_header": bool(fb and fb[0] > 1),
            }
        )
    for name, sheet, coord in _iter_named_regions(wb):
        geo = _region_geometry(wb, sheet, coord)
        if geo is None or geo.get("cardinality") != "multi_cell":
            continue
        out.append(
            {
                "id": region_id_for_name(name),
                "kind": "sample_data",
                "name": name,
                "sheet": sheet,
                "range": coord,
                "rows": geo.get("rows"),
                "cols": geo.get("cols"),
            }
        )
    return out


# ---------------------------------------------------------------------------
# Component inventories (number formats / named styles / tables / CF / charts /
# images) - GEOMETRY/structure evidence only, never a brand-name literal. These
# feed the model's comprehension surface and the validator's survival checks; the
# workbook round-trips these parts byte-intact, so the gap they close is
# comprehension-blindness, not loss.
# ---------------------------------------------------------------------------
def inventory_number_formats(wb) -> list[dict]:
    """Surface the distinct cell number-format masks the workbook actually uses.

    Walks every MATERIALIZED cell (sparse-safe) and aggregates each distinct
    ``number_format`` mask to ``{"format": mask, "count": n, "sample": addr}``,
    sorted by mask. The trivial ``General`` mask is dropped (it carries no brand
    intent). Universal: the masks are the author's OWN formatting, not a code-side
    word list, so this works for any workbook/locale.
    """
    counts: dict[str, int] = {}
    samples: dict[str, str] = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws._cells.values():
            if cell.value is None:
                continue
            fmt = cell.number_format
            if not fmt or fmt == "General":
                continue
            counts[fmt] = counts.get(fmt, 0) + 1
            samples.setdefault(fmt, f"{ws.title}!{cell.coordinate}")
    return [
        {"format": fmt, "count": counts[fmt], "sample": samples[fmt]}
        for fmt in sorted(counts)
    ]


# Currency glyph inside a ``[$X-locale]`` block (X is a real symbol, not the
# bare ``[$-409]`` locale marker) or a bare glyph after brackets/quotes are removed.
_CURRENCY_GLYPHS = "$€£¥₹₩₪฿"


def number_format_family(code: Optional[str]) -> Optional[str]:
    """Classify an OOXML number-format mask into a brand-AGNOSTIC semantic family.

    Maps the author's OWN format mask to one of a small, format-neutral vocabulary
    (``percent`` | ``currency`` | ``accounting`` | ``datetime`` | ``date`` |
    ``time`` | ``decimal`` | ``integer`` | ``text`` | ``scientific``) so a Grid can
    name the *intent* and the profile resolves it to the template's real mask. The
    classification is structural (it reasons over the mask tokens, never a locale
    word list) and conservative: an unrecognized / trivial mask returns ``None`` so
    no role is fabricated. Returns ``None`` for ``General`` / empty.
    """
    c = (code or "").strip()
    if not c or c == "General":
        return None
    low = c.lower()
    # For the date/time/numeric token scan, strip quoted literals AND bracket blocks
    # so stray letters inside them (e.g. "day:" or [Red]) never confuse it.
    bare = re.sub(r'"[^"]*"', "", c)
    bare = re.sub(r"\[[^\]]*\]", "", bare)
    bare_low = bare.lower()
    # For CURRENCY, keep quoted literals (a glyph is often quoted, e.g. ``"$"#,##0``)
    # but drop bracket blocks so the bare ``[$-409]`` LOCALE marker is not read as a
    # currency; an in-bracket symbol ([$€-2], [$$-409]) is detected separately.
    c_nobrackets = re.sub(r"\[[^\]]*\]", "", c)
    has_pct = "%" in bare
    in_bracket_currency = bool(re.search(r"\[\$[^\]\-]", c))  # [$€-2], [$$-409], ...
    has_currency = in_bracket_currency or any(
        g in c_nobrackets for g in _CURRENCY_GLYPHS
    )
    has_date = ("y" in bare_low) or ("d" in bare_low) or ("mmm" in bare_low)
    has_time = ("h" in bare_low) or ("s" in bare_low) or ("a/p" in low)
    has_sci = bool(re.search(r"e[+\-]", bare_low))
    has_text = "@" in bare
    is_numeric = any(ch in bare for ch in "0#?")
    if has_pct:
        return "percent"
    if has_sci:
        return "scientific"
    if has_text and not is_numeric:
        return "text"
    if has_currency:
        # Accounting masks align with parentheses / underscore padding.
        return "accounting" if ("_(" in c or "_)" in c or "(" in bare) else "currency"
    if has_date and has_time:
        return "datetime"
    if has_date:
        return "date"
    if has_time:
        return "time"
    if is_numeric:
        return "decimal" if "." in bare else "integer"
    return None


def number_format_roles(number_formats: list[dict]) -> dict[str, str]:
    """Pick one representative mask per semantic family from the inventory.

    For each ``{"format", "count"}`` entry (from :func:`inventory_number_formats`),
    classify the mask and keep, per family, the mask with the highest in-template
    usage count (first on a tie, since the inventory is mask-sorted -> deterministic).
    Returns ``{"number.<family>": <mask>}`` for the families the template actually
    uses; empty when the workbook carries no classifiable masks.
    """
    best: dict[str, tuple[int, str]] = {}
    for entry in number_formats or []:
        if not isinstance(entry, dict):
            continue
        code = entry.get("format")
        fam = number_format_family(code)
        if not fam:
            continue
        count = int(entry.get("count", 0) or 0)
        current = best.get(fam)
        if current is None or count > current[0]:
            best[fam] = (count, code)
    return {f"number.{fam}": code for fam, (_, code) in best.items()}


def inventory_named_styles(wb) -> list[dict]:
    """Surface each workbook NamedStyle with a structural digest of its props.

    Returns one entry per named cell style::

        {"id": "cell.style.acmetitle", "name": "AcmeTitle",
         "number_format": "...", "font": {...}, "has_fill": bool,
         "has_border": bool, "builtin": bool}

    Evidence only - the id derives from the style's OWN name (slugified), never a
    code word-list, so a ``cell_style`` role can re-assert a present brand style
    on a cover/region fill and the validator can verify it survived. ``builtin``
    flags Excel's reserved 'Normal'/builtin styles so a nomination layer can
    prefer a real brand style over the builtin floor.
    """
    out: list[dict] = []
    for style in getattr(wb, "named_styles", []) or []:
        name = style if isinstance(style, str) else getattr(style, "name", None)
        if not name:
            continue
        entry: dict[str, Any] = {
            "id": f"cell.style.{slugify(name).replace('-', '')}",
            "name": name,
            "builtin": bool(getattr(style, "builtinId", None) is not None)
            or name == "Normal",
        }
        if not isinstance(style, str):
            nf = getattr(style, "number_format", None)
            if nf and nf != "General":
                entry["number_format"] = nf
            font = getattr(style, "font", None)
            if font is not None:
                color = getattr(getattr(font, "color", None), "rgb", None)
                entry["font"] = {
                    "name": getattr(font, "name", None),
                    "size": getattr(font, "size", None),
                    "bold": bool(getattr(font, "bold", False)),
                    "italic": bool(getattr(font, "italic", False)),
                    "color": color if isinstance(color, str) else None,
                }
            fill = getattr(style, "fill", None)
            entry["has_fill"] = bool(
                fill is not None and getattr(fill, "patternType", None)
            )
            border = getattr(style, "border", None)
            entry["has_border"] = bool(
                border is not None
                and any(
                    getattr(getattr(border, side, None), "style", None)
                    for side in ("left", "right", "top", "bottom")
                )
            )
        out.append(entry)
    return out


def inventory_table_styles(wb) -> list[dict]:
    """Surface every native table object's name, ref and applied table style.

    One entry per ``openpyxl`` ``Table`` on any sheet::

        {"name": "AcmeDataTbl", "sheet": "Model", "ref": "A3:G7",
         "style": "TableStyleMedium2", "show_row_stripes": true, ...}

    These round-trip byte-intact; surfacing them lets the model flag a refilled
    region that overlaps a table and lets a survival check assert the table was
    not lost. The style name is the author's data, never matched as a literal.
    """
    out: list[dict] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        try:
            # ``TableList.items()`` yields ``(name, ref_string)``; the Table OBJECTS
            # (which carry ``tableStyleInfo``) are the ``.values()``.
            tables = list(ws.tables.values())
        except Exception:
            continue
        for tbl in tables:
            ref = getattr(tbl, "ref", None)
            entry: dict[str, Any] = {
                "name": str(
                    getattr(tbl, "displayName", "") or getattr(tbl, "name", "")
                ),
                "sheet": sheet,
                "ref": str(ref) if ref else None,
            }
            info = getattr(tbl, "tableStyleInfo", None)
            if info is not None:
                entry["style"] = getattr(info, "name", None)
                entry["show_row_stripes"] = bool(getattr(info, "showRowStripes", False))
                entry["show_col_stripes"] = bool(
                    getattr(info, "showColumnStripes", False)
                )
                entry["show_first_column"] = bool(
                    getattr(info, "showFirstColumn", False)
                )
                entry["show_last_column"] = bool(getattr(info, "showLastColumn", False))
            out.append(entry)
    out.sort(key=lambda e: (e["sheet"], e["name"]))
    return out


def inventory_conditional_formatting(wb) -> list[dict]:
    """Surface each sheet's conditional-formatting rules (sqref + rule types).

    One entry per CF range::

        {"sheet": "Model", "sqref": "B4:E6", "rule_types": ["colorScale"]}

    Lets the model flag a refilled region overlapping a CF range and lets a
    survival check assert CF was preserved. Rule *type* only - no brand literal.
    """
    out: list[dict] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        try:
            cf = ws.conditional_formatting
        except Exception:
            continue
        for rng in cf:
            try:
                sqref = str(getattr(rng, "sqref", rng))
                rules = list(cf[rng])
            except Exception:
                continue
            rule_types = sorted({r.type for r in rules if getattr(r, "type", None)})
            out.append({"sheet": sheet, "sqref": sqref, "rule_types": rule_types})
    out.sort(key=lambda e: (e["sheet"], e["sqref"]))
    return out


def _anchor_cell(anchor) -> Optional[str]:
    """Return the top-left anchor cell ``A1``-coord of a drawing/chart anchor.

    openpyxl two/one-cell anchors expose ``_from`` with 0-based ``col``/``row``;
    a plain string anchor is returned verbatim. None when unresolvable.
    """
    if isinstance(anchor, str):
        return anchor
    frm = getattr(anchor, "_from", None)
    if frm is None:
        return None
    try:
        from openpyxl.utils import get_column_letter

        return f"{get_column_letter(frm.col + 1)}{frm.row + 1}"
    except Exception:
        return None


def inventory_charts(wb) -> list[dict]:
    """Surface each sheet's native charts (type + anchor).

    One entry per chart::

        {"sheet": "Model", "type": "BarChart", "anchor": "I3", "has_title": true}

    A survival check can assert the workbook did not silently drop a chart; the
    chart type is structural metadata, never matched as a literal.
    """
    out: list[dict] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for chart in getattr(ws, "_charts", []) or []:
            entry: dict[str, Any] = {"sheet": sheet, "type": type(chart).__name__}
            cell = _anchor_cell(getattr(chart, "anchor", None))
            if cell is not None:
                entry["anchor"] = cell
            if getattr(chart, "title", None) is not None:
                entry["has_title"] = True
            out.append(entry)
    return out


def inventory_images(wb) -> list[dict]:
    """Surface each sheet's embedded images (drawings).

    One entry per image ``{"sheet": ..., "anchor": <cell>}`` so a survival check
    can assert a logo/picture was not dropped. No brand literal.
    """
    out: list[dict] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for img in getattr(ws, "_images", []) or []:
            entry: dict[str, Any] = {"sheet": sheet}
            cell = _anchor_cell(getattr(img, "anchor", None))
            if cell is not None:
                entry["anchor"] = cell
            out.append(entry)
    return out


# ---------------------------------------------------------------------------
# Ordered skeleton (the profile["structure"] payload; peer of docx)
# ---------------------------------------------------------------------------
def detect_skeleton(wb) -> dict:
    """Detect the workbook's ordered sheet skeleton.

    Returns the ``structure`` section for ``profile.json``::

        {"ordered": True,
         "skeleton": [ {"region","order","role","required","repeatable",...}, ... ]}

    One ``sheet`` region per worksheet in tab order (``ordered`` means the sheet
    order must be respected on generation). The region id matches the
    ``region.sheet.<slug>`` ids from :func:`inventory_regions` so a comprehension
    ``demo``/``required`` attr can be derived onto the matching skeleton region.
    """
    skeleton: list[dict] = []
    for order, sheet in enumerate(wb.sheetnames):
        skeleton.append(
            {
                "region": f"sheet.{slugify(sheet).replace('-', '')}",
                "order": order,
                "role": "section.sheet",
                "required": True,
                "repeatable": False,
                "freeform": False,
                "evidence": f"worksheet {sheet!r} (tab order {order})",
            }
        )
    return {"ordered": True, "skeleton": skeleton}
