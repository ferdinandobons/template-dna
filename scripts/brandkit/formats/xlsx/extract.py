# SPDX-License-Identifier: MIT
"""XLSX extraction.

Emits the format-uniform comprehension inventories (plan §4) from GEOMETRY
evidence, with **no** range-name literals. The old ``"title_cell"`` /
``"data_region"`` special-casing is gone: every named range is surfaced
generically with its geometry (cardinality / merged header / frozen band / table
membership) into ``surface.xlsx.cover_anchors``; multi-cell named ranges are
sample-data candidates in ``surface.xlsx.regions``; ``surface.xlsx.fields`` is the
legal-empty field inventory (a workbook has no TOC-style field code). Purpose is
the model's job (``cover_slots`` / ``demo_classification``); the extractor only
records what the geometry proves.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from brandkit.formats import catalog
from brandkit.formats.xlsx import structure as xlsx_structure
from brandkit.ooxml import pack
from brandkit.profile import schema, store


def extract(
    template: str | Path,
    name: str,
    *,
    scope: str = "project",
    cwd: str | Path | None = None,
) -> Path:
    template_path = Path(template)
    wb = load_workbook(template_path, data_only=False)

    named_regions = xlsx_structure.named_regions_map(wb)
    # Format-uniform inventories the model reasons over and the validator binds to
    # (plan §4). Geometry evidence only - no range-name literals. ``fields`` is the
    # legal-empty xlsx field inventory; ``cover_slots``/``conventions.indexes`` for
    # xlsx therefore have nothing to bind to (readiness gate), and the in-scope
    # comprehension surface is ``demo_classification`` keyed to ``regions``.
    cover_anchors = xlsx_structure.inventory_cover_anchors(wb)
    fields = xlsx_structure.inventory_fields(wb)
    regions = xlsx_structure.inventory_regions(wb)
    skeleton = xlsx_structure.detect_skeleton(wb)

    named_styles = xlsx_structure.inventory_named_styles(wb)
    number_formats = xlsx_structure.inventory_number_formats(wb)
    table_styles = xlsx_structure.inventory_table_styles(wb)
    conditional_formatting = xlsx_structure.inventory_conditional_formatting(wb)
    charts = xlsx_structure.inventory_charts(wb)
    images = xlsx_structure.inventory_images(wb)

    roles = _roles(named_regions, named_styles, number_formats)
    surface = {
        "xlsx": {
            "sheets": wb.sheetnames,
            "named_regions": named_regions,
            "named_styles": named_styles,
            "number_formats": number_formats,
            "table_styles": table_styles,
            "conditional_formatting": conditional_formatting,
            "charts": charts,
            "images": images,
            "cover_anchors": cover_anchors,
            "fields": fields,
            "regions": regions,
        }
    }
    profile = schema.build_envelope(
        "xlsx",
        {"name": name, "display_name": name},
        extracted_at=datetime.now(timezone.utc).isoformat(),
        source_template={
            "filename": template_path.name,
            "sha256": store.sha256_file(template_path),
        },
        theme=_theme(),
        roles=roles,
        surface=surface,
        structure=skeleton,
    )
    # Anchors summary derived from GEOMETRY, never from a literal range name. The
    # cover anchor is "present" when any named region exists to host a cover slot;
    # the demo region is "present" when any multi-cell (sample-data) region exists.
    has_sample_data = any(r.get("kind") == "sample_data" for r in regions)
    profile["anchors"] = {
        "cover": {
            "kind": schema.AnchorKind.NAMED_RANGE.value
            if cover_anchors
            else schema.AnchorKind.NONE.value,
            "slots_found": len(cover_anchors),
        },
        "demo_region": {"present": has_sample_data},
        "toc": {"present": False},
    }
    profile["provenance"]["ooxml_parts_seen"] = pack.list_parts(template_path)
    profile["artifact_catalog"] = _artifact_catalog(
        template_path, wb, named_regions, profile["provenance"]["ooxml_parts_seen"]
    )
    profile["capabilities"] = _capabilities()
    target = store.target_dir_for_save(name, scope, cwd=cwd)
    return store.save_profile(
        target,
        profile,
        template_path.read_bytes(),
        extra_files={"PROFILE.md": _profile_md(profile)},
        overwrite=True,
    )


def _roles(
    named_regions: dict,
    named_styles: list[dict] | None = None,
    number_formats: list[dict] | None = None,
) -> dict:
    """Build the role registry from the named ranges + named cell styles (no literals).

    Every named range becomes a ``named_range`` role keyed by its slugified name;
    none is privileged by a hardcoded name. Whether a region is a cover title, a
    data block, or something else is the comprehension model's call
    (``cover_slots`` / ``demo_classification``), never a code-side word.

    Additionally, every PRESENT non-builtin brand cell style is promoted into a
    ``cell_style`` role with a VERBATIM resolver target (its own style name), so a
    cover/region fill can re-assert the brand style and the validator can verify it
    survived - a structural nomination (the style is in the workbook), never a name
    lexicon match (CC-2 / X3). Builtin styles (Normal/...) are not nominated.

    Finally, the workbook's OWN distinct number-format masks are promoted into
    ``number.<family>`` roles (``number_format`` resolver). Each is a semantic
    family (currency/percent/date/...) bound to the verbatim mask the template
    uses, so a Grid can name the *intent* and the resolver fills the real mask -
    brand-by-construction, never a fabricated format. Classification is structural
    (mask tokens), best-effort.

    When the workbook declares no named range AND no brand style at all, a single
    default ``cell_style`` role keeps the registry non-empty so generation has a
    fallback.
    """
    roles = {"_index": []}

    def add(
        rid: str,
        resolver: dict,
        signal: str,
        *,
        status: str = schema.Status.ROBUST.value,
        confidence: float = 0.85,
    ) -> None:
        roles[rid] = {
            "resolver": resolver,
            "appearance": {},
            "verified": status != schema.Status.STUB.value,
            "confidence": confidence,
            "status": status,
            "evidence": {"signal": signal},
        }
        roles["_index"].append(rid)

    # Region role ids derive from the SAME helper the structure inventory uses, so
    # a role id and its inventory id can never drift (the binding depends on them
    # agreeing) - Q7.
    for name in sorted(named_regions):
        rid = xlsx_structure.region_id_for_name(name)
        add(
            rid,
            {"type": schema.ResolverType.NAMED_RANGE.value, "name": name},
            f"named range {name}",
        )

    # Promote present brand cell styles into cell_style roles (structural; the id +
    # resolver target are the style's OWN name). The inventory already computes the
    # canonical id and flags builtins, so reuse it (no inline slug duplication).
    for style in sorted(named_styles or [], key=lambda s: s.get("name") or ""):
        name = style.get("name")
        if not name or style.get("builtin"):
            continue
        rid = style.get("id")
        if not rid:
            continue
        add(
            rid,
            {"type": schema.ResolverType.CELL_STYLE.value, "style_name": name},
            f"named cell style {name}",
        )

    # Promote the workbook's distinct number-format masks into number.<family>
    # roles (number_format resolver). The family is the brand-agnostic intent a
    # Grid names; the resolver target is the template's own verbatim mask.
    for rid, mask in sorted(xlsx_structure.number_format_roles(number_formats).items()):
        add(
            rid,
            {"type": schema.ResolverType.NUMBER_FORMAT.value, "number_format": mask},
            f"number format mask {mask!r}",
            status=schema.Status.BEST_EFFORT.value,
            confidence=0.7,
        )

    if not roles["_index"]:
        add(
            "cell.default",
            {"type": schema.ResolverType.CELL_STYLE.value, "style_name": "Normal"},
            "default style",
        )
    return roles


def _artifact_catalog(path: Path, wb, named_regions: dict, parts: list[str]) -> dict:
    out = catalog.part_catalog(path)
    out["ooxml_parts"] = parts
    out["named_ranges"] = named_regions
    out["named_styles"] = [
        style if isinstance(style, str) else style.name for style in wb.named_styles
    ]
    out["sheets"] = {}
    formulas = {}
    for ws in wb.worksheets:
        sheet_info = {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "tables": list(ws.tables.keys()),
            "merged_cells": [str(rng) for rng in ws.merged_cells.ranges],
            "dimensions": {
                "column_widths": {
                    key: dim.width
                    for key, dim in ws.column_dimensions.items()
                    if dim.width
                },
                "row_heights": {
                    str(key): dim.height
                    for key, dim in ws.row_dimensions.items()
                    if dim.height
                },
            },
            "non_empty_cells": [],
        }
        # Iterate materialized cells only; large corporate models often have
        # broad dimensions with sparse content.
        for cell in ws._cells.values():
            if cell.value is None:
                continue
            address = f"{ws.title}!{cell.coordinate}"
            value = cell.value
            is_formula = isinstance(value, str) and value.startswith("=")
            if is_formula:
                formulas[address] = value
            entry = {
                "address": address,
                "data_type": cell.data_type,
                "style": cell.style,
                "number_format": cell.number_format,
            }
            # Carry the cell's literal TEXT (only non-formula strings) so the shared
            # comprehension excerpt collector has real text to sample for xlsx -
            # without it the model would reason over geometry alone (CC-1). A
            # formula / number / date is not surfaced as excerpt text (the formula
            # catalog already records formulas; numerics carry no brand wording).
            if isinstance(value, str) and not is_formula:
                entry["text"] = value
            sheet_info["non_empty_cells"].append(entry)
        out["sheets"][ws.title] = sheet_info
    out["formulas"] = formulas
    # Native-component baselines for the QA component-survival check: the workbook
    # round-trips these byte-intact, so a mismatch shell-vs-output is a regression.
    out["tables"] = xlsx_structure.inventory_table_styles(wb)
    out["conditional_formatting"] = xlsx_structure.inventory_conditional_formatting(wb)
    out["charts"] = xlsx_structure.inventory_charts(wb)
    out["images"] = xlsx_structure.inventory_images(wb)
    return out


def _capabilities() -> dict:
    return {
        "extracts_all_ooxml_parts": True,
        "extracts_named_ranges": True,
        "extracts_formula_catalog": True,
        "preserves_formulas_in_shell": True,
        "region_bounds_guard": True,
        "generates_from_shell": True,
        "emits_region_geometry": True,
        "comprehension_demo_classification": True,
        "native_charts": True,
        "resolves_number_formats": True,
    }


def _theme() -> dict:
    return {
        "colors": {},
        "palette_roles": {"primary": {"theme": "accent1"}, "text": {"theme": "dk1"}},
        "fonts": {
            "major": {"latin": None, "fallback": "Arial"},
            "minor": {"latin": None, "fallback": "Calibri"},
        },
        "embedded_fonts": [],
    }


def _profile_md(profile: dict) -> str:
    return (
        "# Brand Profile: " + profile["identity"]["display_name"] + "\n\n- kind: xlsx\n"
    )
