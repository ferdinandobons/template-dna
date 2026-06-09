# SPDX-License-Identifier: MIT
"""XLSX generation from a GridDocument.

Routed through the SHARED resolver spine (``ProfileResolver``): every named-cell /
named-region fill is resolved to its ``named_range`` resolver op by the same
kind-aware spine the docx/pptx generators use, so the brand guarantee (only fill
a region the profile PROVED exists, with a resolver type legal for ``xlsx``) is
enforced in one place. Formulas are NEVER authored AND never overwritten - both
fill loops skip a target cell that already holds a shell formula (the same guard
``_clear_region`` uses), so refilling a region that straddles formula cells
preserves them verbatim; the workbook is then marked for a full recalc on open so
those preserved formulas pick up the new inputs. When
comprehension is present it steers the cover (fill / clear / leave named regions)
and demo (clear sample-data regions ruled ``verdict=demo``); when absent the
deterministic path fills exactly the named cells/regions the grid names.
"""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.chart import (
    AreaChart,
    BarChart,
    DoughnutChart,
    LineChart,
    PieChart,
    Reference,
)
from openpyxl.utils.cell import range_boundaries

from brandkit.grid.model import GridDocument
from brandkit.ooxml.idempotency import repack_fixed_timestamps
from brandkit.profile import schema, store
from brandkit.profile.reconcile import confidence_clears_floor
from brandkit.profile.resolver import ProfileResolver
from brandkit.qa.checks_deterministic import check_no_net_structure_loss
from brandkit.qa.model import Finding


def generate(
    profile: dict,
    shell_path: str | Path,
    grid: GridDocument,
    output: str | Path,
    *,
    findings: Optional[list[Finding]] = None,
) -> Path:
    """Generate an XLSX from ``grid`` onto the brand ``shell_path``.

    ``findings`` (optional out-param) is appended with any degradation findings so
    the QA gate can surface them. Off-brand / fabricated fills stay impossible by
    construction: every fill target is resolved through :class:`ProfileResolver`
    against ``profile['roles']``, so a grid key that does not name a region the
    profile proved exists is rejected (never invented).
    """
    sink: list[Finding] = findings if findings is not None else []
    wb = load_workbook(shell_path, data_only=False)
    resolver = ProfileResolver(profile)

    # The named-range geometry the resolver's ``named_range`` ops point at. This is
    # the profile's OWN surfaced map (the author's range names), not a code literal.
    regions = ((profile.get("surface") or {}).get("xlsx") or {}).get(
        "named_regions"
    ) or {}

    # Reverse index: named-range NAME -> the role id whose resolver targets it. The
    # grid is keyed by the author's range names (its vocabulary); we resolve each
    # through the spine so the brand guarantee gates every fill.
    name_to_role = _name_to_role(profile)

    # Cover anchors whose brand cell STYLE must be re-asserted after fill (X4):
    # the named-range NAME -> the originally-applied brand named style id. A
    # value-only write preserves the style in openpyxl, but a clear-then-refill or
    # a fill onto a re-armed placeholder can drop it; re-asserting the captured
    # style (a verbatim id read off the shell, never a literal) keeps a filled
    # cover cell guaranteed brand-styled rather than incidentally so.
    cover_style_for_name = _cover_anchor_styles(wb, profile, regions)

    # Fill single named cells. A target cell that already holds a shell formula
    # is NEVER overwritten (mirrors ``_clear_region``'s guard): a named single
    # cell can legally point at a formula output, and silently clobbering it is
    # data loss the QA formula-preservation check would (now) flag, but the guard
    # keeps it impossible by construction. ``None`` values are skipped so a sparse
    # cell write never blanks a preserved cell.
    for name, value in grid.cells.items():
        target = _resolve_named_target(resolver, name_to_role, regions, name)
        sheet, coord = target["sheet"], target["range"]
        min_col, min_row, _, _ = range_boundaries(coord)
        cell = wb[sheet].cell(row=min_row, column=min_col)
        wrote = _fill_cell(cell, value, sink=sink, where=name)
        _reassert_cover_style(cell, cover_style_for_name.get(name))
        # Apply the resolved number-format intent LAST so it wins over any mask the
        # re-asserted cell style carries (the explicit semantic intent is authoritative)
        # - but ONLY when a value was actually written, so a preserved-formula / merged
        # cell never has its own format silently clobbered.
        if wrote:
            mask = _resolve_number_mask(resolver, grid.formats.get(name), name, sink)
            if mask:
                cell.number_format = mask

    # Fill multi-cell named regions (bounds-guarded; never overruns the range).
    # Same formula guard as the single-cell loop: a region named over a block that
    # straddles formula cells (e.g. a body range whose trailing columns are SUM/IF
    # totals) must refill only its literal cells, never erase the formulas.
    for name, values in grid.regions.items():
        target = _resolve_named_target(resolver, name_to_role, regions, name)
        sheet, coord = target["sheet"], target["range"]
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        _check_region_bounds(
            name, values, max_rows=max_row - min_row + 1, max_cols=max_col - min_col + 1
        )
        ws = wb[sheet]
        # Resolve the region's number-format intent ONCE (one degrade at most), then
        # apply the same mask to each filled cell.
        region_mask = _resolve_number_mask(resolver, grid.formats.get(name), name, sink)
        for r_idx, row in enumerate(values):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=min_row + r_idx, column=min_col + c_idx)
                wrote = _fill_cell(cell, value, sink=sink, where=name)
                # Only format cells we actually wrote: never clobber a preserved
                # formula cell's (or a merged slave's) own format.
                if wrote and region_mask:
                    cell.number_format = region_mask

    # Native charts over the workbook's OWN cell data (after the fills, so the
    # referenced ranges carry the new values). The chart inherits the workbook
    # theme's accent colors - on-brand by construction, no literal color written.
    for spec in grid.charts:
        _write_chart(wb, spec, sink)

    # Comprehension-steered reconciliation (no-op when comprehension is absent):
    # CLEAR cover anchors / demo sample-data regions the model ruled destructive,
    # with the destructive floor corroborating each removal.
    removed_refs = _reconcile_cover_and_demo(wb, profile, grid, regions, sink)

    # Recalc: mark the workbook for a full recompute on open so preserved formulas
    # pick up the new inputs. The analogue of the docx TOC refresh; formulas are
    # NEVER authored here, only the shell's own formulas are recalculated. Narrow
    # except: an openpyxl API change must surface (a workbook that never recomputes
    # preserved formulas would silently ship), so only the expected attribute miss
    # is tolerated.
    try:
        wb.calculation.fullCalcOnLoad = True
    except AttributeError:
        sink.append(
            Finding(
                check="xlsx_recalc",
                severity=schema.Severity.WARNING.value,
                message="could not set fullCalcOnLoad; preserved formulas may not recompute on open",
            )
        )

    # Destructive-action floor (plan §6): every region the reconciliation cleared
    # must carry a corroborated destructive verdict AND clear the confidence floor,
    # else ERROR. Model-free. The confidence threaded here is the model's single
    # comprehension confidence (the same value the reconcile site gates on).
    if store.comprehension_is_present(profile):
        comp = profile.get("comprehension")
        confidence = (
            float(comp.get("confidence") or 0.0) if isinstance(comp, dict) else None
        )
        sink.extend(
            check_no_net_structure_loss(removed_refs, profile, confidence=confidence)
        )

    out = Path(output)
    out.parent.mkdir(parents=True, exist_ok=True)
    # Byte-idempotency precondition: when the shell's core.xml carries no
    # dcterms:created, openpyxl FABRICATES a wall-clock created (and stamps it into
    # modified) at load, which two generations straddling a 2s boundary would differ
    # on - and pinning modified-from-created cannot fix it because created is itself
    # wall-clock. Pin a fixed created in that case so the output is deterministic.
    # (When the shell HAS a created, openpyxl preserved it, so leave it untouched.)
    if _shell_created_iso(shell_path) is None:
        wb.properties.created = datetime(1980, 1, 1, tzinfo=timezone.utc)
    wb.save(out)
    # Byte-idempotency (X7): openpyxl's writer stamps the current wall-clock time
    # into every ZIP entry header AND into ``docProps/core.xml`` ``dcterms:modified``
    # at save (overriding ``wb.properties.modified``), so two otherwise-identical
    # generations differ only by the save second. Normalize the saved package
    # post-hoc - fixed ZIP timestamps + a ``modified`` pinned to the package's own
    # ``created`` - so re-running the generator yields an identical file. No
    # code-literal date is invented.
    repack_fixed_timestamps(out, pin_modified_from_created=True)
    return out


def _new_xlsx_chart(chart_type: str | None):
    """Map an IR ``chart_type`` to a fresh openpyxl chart object, or None if unknown.

    ``bar`` is the common business vertical bar (a COLUMN chart); ``barh`` is the
    true horizontal bar. No explicit series color is set, so the chart inherits the
    workbook theme's accent palette."""
    kind = (chart_type or "bar").lower()
    if kind in ("bar", "column"):
        chart = BarChart()
        chart.type = "col"
        return chart
    if kind == "barh":
        chart = BarChart()
        chart.type = "bar"
        return chart
    if kind == "line":
        return LineChart()
    if kind == "area":
        return AreaChart()
    if kind == "pie":
        return PieChart()
    if kind == "doughnut":
        return DoughnutChart()
    return None


def _write_chart(wb, spec: dict, sink: list) -> None:
    """Author a NATIVE xlsx chart that REFERENCES cell data (the xlsx peer of the
    docx/pptx inline-data chart - here the data lives in the sheet, the chart's
    strength). The chart inherits the workbook theme's accent colors (on-brand by
    construction; no literal color). A spec missing its anchor/data, naming a sheet
    not in the workbook, or carrying a malformed range, degrades to a loud
    ``block_degraded`` WARNING - never a crash. An unknown ``type`` falls back to a
    clustered column chart (INFO)."""
    anchor = spec.get("anchor")
    data_ref = spec.get("data")
    if not anchor or not data_ref:
        sink.append(
            Finding(
                check="block_degraded",
                severity=schema.Severity.WARNING.value,
                message="'chart' spec missing anchor/data; skipped",
            )
        )
        return
    sheet = spec.get("sheet")
    if sheet and sheet not in wb.sheetnames:
        sink.append(
            Finding(
                check="block_degraded",
                severity=schema.Severity.WARNING.value,
                message=f"'chart' references sheet {sheet!r} not in the workbook; skipped",
            )
        )
        return
    ws = wb[sheet] if sheet else wb.active

    chart = _new_xlsx_chart(spec.get("type"))
    if chart is None:
        chart = BarChart()
        chart.type = "col"
        sink.append(
            Finding(
                check="chart_type_fallback",
                severity=schema.Severity.INFO.value,
                message=f"chart_type {spec.get('type')!r} unknown; "
                "rendered as a clustered column chart",
            )
        )

    try:
        d_min_c, d_min_r, d_max_c, d_max_r = range_boundaries(str(data_ref))
        chart.add_data(
            Reference(
                ws, min_col=d_min_c, min_row=d_min_r, max_col=d_max_c, max_row=d_max_r
            ),
            titles_from_data=bool(spec.get("data_titles", True)),
        )
        cats = spec.get("categories")
        if cats:
            c_min_c, c_min_r, c_max_c, c_max_r = range_boundaries(str(cats))
            chart.set_categories(
                Reference(
                    ws,
                    min_col=c_min_c,
                    min_row=c_min_r,
                    max_col=c_max_c,
                    max_row=c_max_r,
                )
            )
        if spec.get("title"):
            chart.title = str(spec["title"])
        ws.add_chart(chart, str(anchor))
    except Exception as exc:
        sink.append(
            Finding(
                check="block_degraded",
                severity=schema.Severity.WARNING.value,
                message=f"'chart' not placed ({exc}); skipped",
            )
        )


def _name_to_role(profile: dict) -> dict[str, str]:
    """Map each named-range NAME to the role id whose resolver targets it.

    Built from ``profile['roles']`` so a grid key is resolved through the spine
    (the brand guarantee) rather than read straight off ``surface``. A role whose
    resolver is not a ``named_range`` (e.g. the default ``cell_style`` fallback) is
    skipped - it names no range.
    """
    out: dict[str, str] = {}
    for rid, entry in (profile.get("roles") or {}).items():
        if rid == "_index" or not isinstance(entry, dict):
            continue
        resolver = entry.get("resolver") or {}
        if resolver.get("type") != schema.ResolverType.NAMED_RANGE.value:
            continue
        name = resolver.get("name")
        if isinstance(name, str) and name:
            out[name] = rid
    return out


def _resolve_named_target(resolver, name_to_role, regions, name) -> dict:
    """Resolve a grid key (a named-range name) to its ``{sheet, range}`` target.

    Routes through the shared spine: the name maps to a role id, the resolver op
    must be a legal ``named_range`` for the profile kind, and its target range must
    be one the profile surfaced. A key naming no proven region is an ERROR (the
    fill is never invented).
    """
    rid = name_to_role.get(name)
    if rid is not None:
        op = resolver.resolve_role(rid, fallback=None)
        if op.resolver.get("type") == schema.ResolverType.NAMED_RANGE.value:
            resolved_name = op.resolver.get("name", name)
            target = regions.get(resolved_name)
            if target:
                return target
    # Fail closed: a grid key that resolves to no surfaced named range is a brand
    # breach (a fabricated target), not a silent no-op.
    raise ValueError(f"unknown named cell/range {name!r}")


def _reconcile_cover_and_demo(
    wb, profile: dict, grid: GridDocument, regions: dict, findings: list[Finding]
) -> set[str]:
    """Reconcile preserved cover anchors + demo sample-data regions with new content.

    Comprehension-steered; a no-op when comprehension is absent (the deterministic
    fills above already wrote the named regions the grid named). For each
    ``cover_slots`` entry ruled ``fill_rule=clear`` and each ``demo_classification``
    region ruled ``verdict=demo`` that maps to a surfaced named range, the cells are
    CLEARED in place - but ONLY when the model's ``comprehension.confidence`` clears
    the destructive floor; otherwise the clear is downgraded to KEEP + WARNING (a
    wrong delete is unrecoverable). The confidence gate mirrors the docx/pptx cover
    reconcilers so the SAME confidence yields the SAME behavior across formats.
    Returns the set of refs actually cleared (anchor_ref / region_ref) for the
    destructive floor. Formulas are never cleared - only the region's own input cells.
    """
    if not store.comprehension_is_present(profile):
        return set()
    comp = profile.get("comprehension")
    if not isinstance(comp, dict):
        return set()

    # The model's single comprehension confidence - the SAME value the cover
    # reconcilers gate on. Below the floor every destructive clear is downgraded.
    confidence = float(comp.get("confidence") or 0.0)
    floor_cleared = confidence_clears_floor(confidence)

    cover_anchors = ((profile.get("surface") or {}).get("xlsx") or {}).get(
        "cover_anchors"
    ) or []
    region_inv = ((profile.get("surface") or {}).get("xlsx") or {}).get("regions") or []
    anchor_to_name = {
        a.get("id"): a.get("name") for a in cover_anchors if isinstance(a, dict)
    }
    region_to_name = {
        r.get("id"): r.get("name") for r in region_inv if isinstance(r, dict)
    }

    removed: set[str] = set()

    # Cover anchors ruled CLEAR.
    for anchor_ref, slot in (comp.get("cover_slots") or {}).items():
        if (
            not isinstance(slot, dict)
            or slot.get("fill_rule") != schema.FillRule.CLEAR.value
        ):
            continue
        name = anchor_to_name.get(anchor_ref)
        target = regions.get(name) if name else None
        if not target:
            continue
        if not floor_cleared:
            findings.append(
                Finding(
                    check="cover_clear_downgraded",
                    severity=schema.Severity.WARNING.value,
                    message=(
                        f"cover slot {anchor_ref!r} clear not corroborated "
                        f"(confidence {confidence:.2f}); kept"
                    ),
                )
            )
            continue
        if _clear_region(wb, target):
            removed.add(anchor_ref)

    # Demo sample-data regions ruled DEMO -> clear in place (the demo rows go away;
    # the new content is written into the same named region by the grid above).
    for reg in (comp.get("demo_classification") or {}).get("regions") or []:
        if not isinstance(reg, dict) or reg.get("verdict") != schema.Verdict.DEMO.value:
            continue
        region_ref = reg.get("region_ref")
        name = region_to_name.get(region_ref)
        target = regions.get(name) if name else None
        if not target:
            continue
        if not floor_cleared:
            findings.append(
                Finding(
                    check="demo_clear_downgraded",
                    severity=schema.Severity.WARNING.value,
                    message=(
                        f"demo region {region_ref!r} clear not corroborated "
                        f"(confidence {confidence:.2f}); kept"
                    ),
                )
            )
            continue
        # If the grid already refilled this region, the demo rows are overwritten;
        # only clear the trailing cells the new content did NOT cover so no stale
        # demo value survives. Still counts as a destructive act for the floor.
        if _clear_region(wb, target, skip_rows=len(grid.regions.get(name) or [])):
            removed.add(region_ref)

    return removed


def _cover_anchor_styles(wb, profile: dict, regions: dict) -> dict[str, str]:
    """Map each cover-anchor named-range NAME -> its original brand named-style id.

    Captured BEFORE any fill/clear so a re-arm or clear-then-refill cannot lose
    the brand style. Only non-builtin named styles are captured (a builtin
    'Normal' carries no brand intent); the value is the cell's OWN style id read
    off the shell, never a code literal. Empty when the surface has no cover
    anchors (the deterministic-only path preserves style automatically anyway).
    """
    anchors = ((profile.get("surface") or {}).get("xlsx") or {}).get(
        "cover_anchors"
    ) or []
    builtins = {"Normal", ""}
    out: dict[str, str] = {}
    for a in anchors:
        if not isinstance(a, dict):
            continue
        name = a.get("name")
        target = regions.get(name) if name else None
        if not target:
            continue
        sheet, coord = target.get("sheet"), target.get("range")
        if not sheet or not coord or sheet not in wb.sheetnames:
            continue
        try:
            min_col, min_row, _, _ = range_boundaries(coord)
        except Exception:
            continue
        style = wb[sheet].cell(row=min_row, column=min_col).style
        if isinstance(style, str) and style not in builtins:
            out[name] = style
    return out


def _reassert_cover_style(cell, style_id: Optional[str]) -> None:
    """Re-apply ``style_id`` to ``cell`` after a fill, if it changed (X4).

    A no-op when no brand style was captured for the anchor, the cell already
    carries it, or the target is a merged-range slave (read-only in openpyxl);
    otherwise the captured named style (a verbatim shell id) is re-asserted so a
    filled cover slot is guaranteed brand-styled.
    """
    if style_id and not isinstance(cell, MergedCell) and cell.style != style_id:
        cell.style = style_id


def _holds_formula(cell) -> bool:
    """Return True if ``cell`` already holds a shell formula (``=...``).

    The single source of truth for the formula guard shared by the fill loops and
    :func:`_clear_region`: a formula is load-bearing shell content that the
    generator never authors and must never overwrite or blank.
    """
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _shell_created_iso(shell_path) -> str | None:
    """Return the shell core.xml ``dcterms:created`` ISO string, or ``None`` if absent.

    Used only to decide whether openpyxl will fabricate a wall-clock ``created`` on
    save (it does when the shell has none) so the xlsx generator can pin a
    deterministic value and stay byte-idempotent. Any read error is treated as
    "absent" (the conservative, determinism-preserving choice).
    """
    import zipfile

    try:
        with zipfile.ZipFile(shell_path) as zf:
            core = zf.read("docProps/core.xml").decode("utf-8", errors="replace")
    except (OSError, KeyError, zipfile.BadZipFile):
        return None
    match = re.search(r"<dcterms:created[^>]*>([^<]*)</dcterms:created>", core)
    value = match.group(1).strip() if match else ""
    return value or None


def _resolve_number_mask(
    resolver: ProfileResolver, family: Optional[str], where: str, sink: list[Finding]
) -> Optional[str]:
    """Resolve a semantic number ``family`` to the template's own mask, or ``None``.

    Routes ``number.<family>`` through the shared resolver spine. Returns the mask
    when the profile carries that ``number_format`` role; otherwise degrades loudly
    (one WARNING) and returns ``None`` so the caller leaves the cell's existing
    format - a format is NEVER fabricated (brand-by-construction, fail-closed).
    """
    if not family:
        return None
    op = resolver.resolve_role(f"number.{family}", fallback=None)
    if op.resolver_type == schema.ResolverType.NUMBER_FORMAT.value:
        mask = op.resolver.get("number_format")
        if mask:
            return str(mask)
    sink.append(
        Finding(
            check="number_format_degraded",
            severity=schema.Severity.WARNING.value,
            message=(
                f"number format {family!r} for {where!r} is not defined in this "
                f"template (cell left with its existing format)"
            ),
        )
    )
    return None


def _fill_cell(
    cell, value, *, sink: Optional[list[Finding]] = None, where: str | None = None
) -> bool:
    """Write ``value`` into ``cell`` unless that would destroy or escape shell state.

    Returns ``True`` iff a value was actually written (so the caller knows whether a
    number-format mask may be applied without clobbering a preserved/merged cell).

    Skips the write (returns ``False``) when ``value`` is ``None`` (a sparse / ragged
    grid row must not blank a preserved cell), when the target is a merged-range SLAVE
    cell (only a merge's top-left anchor is writable in openpyxl; writing a slave
    raises ``AttributeError``), or when the target already holds a formula (preserving
    the shell's load-bearing formula verbatim). Every other value - including the
    empty string the grid may use for an intentionally cleared cell - is written.

    The engine NEVER authors formulas (they live only in the shell). An author value
    that is a string starting with ``=`` would become a live formula in openpyxl
    (``data_type 'f'``), breaking that invariant and shipping a formula-injection
    payload (``=WEBSERVICE``/``=HYPERLINK``/DDE) in an otherwise on-brand workbook. It
    is neutralized to a TEXT cell (written verbatim, never executed) and surfaced
    loudly, so the content is visible and the neutralization is auditable.

    A dropped non-``None`` value on a merged slave is surfaced as a
    ``block_degraded`` WARNING on ``sink`` (the merged banner keeps its shell value)
    so the skip is visible in QA rather than a silent loss.
    """
    if value is None:
        return False
    if isinstance(cell, MergedCell):
        if sink is not None:
            loc = f"{where} ({cell.coordinate})" if where else cell.coordinate
            sink.append(
                Finding(
                    check="block_degraded",
                    severity=schema.Severity.WARNING.value,
                    message=(
                        "value not written to merged-region slave cell "
                        f"{loc} (only the merge anchor is writable); the merged "
                        "banner kept its shell value"
                    ),
                    location=loc,
                )
            )
        return False
    if _holds_formula(cell):
        return False
    if isinstance(value, str) and value.startswith("="):
        # Neutralize: write verbatim but force a string cell so Excel never executes
        # it as a formula (the engine does not author formulas).
        cell.value = value
        cell.data_type = "s"
        if sink is not None:
            loc = f"{where} ({cell.coordinate})" if where else cell.coordinate
            sink.append(
                Finding(
                    check="formula_injection_neutralized",
                    severity=schema.Severity.WARNING.value,
                    message=(
                        f"author value starting with '=' written as TEXT at {loc}: the "
                        "engine never authors formulas, and a live formula here would "
                        "be a formula-injection risk in the shipped workbook"
                    ),
                    location=loc,
                )
            )
        return True
    cell.value = value
    return True


def _clear_region(wb, target: dict, *, skip_rows: int = 0) -> bool:
    """Empty a named region's NON-FORMULA cells in place. Returns True if it ran.

    A cell holding a formula (``=...``) is never cleared (formulas are load-bearing
    shell content); only literal demo/placeholder cells are emptied. ``skip_rows``
    leaves the first N region rows intact (already refilled by the grid).
    """
    sheet, coord = target.get("sheet"), target.get("range")
    if not sheet or not coord or sheet not in wb.sheetnames:
        return False
    try:
        min_col, min_row, max_col, max_row = range_boundaries(coord)
    except Exception:
        return False
    ws = wb[sheet]
    for r in range(min_row + skip_rows, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue  # a merged slave is read-only; clearing it is a no-op by
                # design (the merge anchor carries the value, slaves inherit it)
            if _holds_formula(cell):
                continue  # preserve formulas
            cell.value = None
    return True


def _check_region_bounds(
    name: str, values: list[list], *, max_rows: int, max_cols: int
) -> None:
    if len(values) > max_rows:
        raise ValueError(
            f"region {name!r} has {len(values)} rows; named range allows {max_rows}"
        )
    for idx, row in enumerate(values, start=1):
        if len(row) > max_cols:
            raise ValueError(
                f"region {name!r} row {idx} has {len(row)} columns; named range allows {max_cols}"
            )
